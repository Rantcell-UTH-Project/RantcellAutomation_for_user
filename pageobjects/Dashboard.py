import re, allure
import time
import pandas as pd
from module_controllers.module_controllers import device_custom_query_module_controllers, \
    date_and_time_module_controllers, operatorcomparison_vs_pdf_module_controllers, floorplan_module_controllers, \
    litetestdata_module_controllers, protestdata_module_controllers, exports_module_controllers
from utils.commonutilis import wait_for_all_elements_presence , find_element
from locators.locators import dashbord_more_info_components
import json
import re, allure
import pandas as pd
import pytest
from utils.createFolderforRantcell_automation_DataandReports import create_folder_for_downloads
from utils.readexcel import *
from pageobjects.login_logout import *
from openpyxl.utils.dataframe import dataframe_to_rows
import concurrent.futures
from configurations.config import ReadConfig as config
import queue
from decimal import Decimal, ROUND_HALF_UP
import win32com.client as win32
import calendar
from locators.locators import *

def side_menu_Components_(driver, device, campaign, userid, password,excelpath):
    try:
        Variable_MobileDevice_Xpath = (By.XPATH, f"//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        protestdata_runvalue = protestdata_module_controllers()
        litetestdata_runvalue = litetestdata_module_controllers()
        if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='Pro TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='LITE TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        print(Variable_MobileDevice_Xpath)
        classifier = (By.XPATH, "//span[normalize-space()='"+str(campaign)+"']", str(campaign))
        click_on_side_bar_menu_compnents(driver, device, userid, password, Variable_MobileDevice_Xpath, classifier)
        count_find_the_campaign = searching_visibility_of_campaigns_by_driver_refresh(driver, classifier, device, userid, password,Variable_MobileDevice_Xpath)
        click_on_campaigns(driver, device, classifier, campaign, count_find_the_campaign, excelpath)
    except Exception as e:
        with allure.step("Failed in side bar menu"):
            allure.attach(driver.get_screenshot_as_png(), name=f"Failed in side bar menu",attachment_type=allure.attachment_type.PNG)
            pytest.fail(str(e))
def searching_visibility_of_campaigns_by_driver_refresh(driver,classifier,device, userid, password,Variable_MobileDevice_Xpath):
    action = ActionChains(driver)
    count_find_the_campaign = 0
    try:
        classifier_elements= driver.find_elements(classifier[0],classifier[1])
        if len(classifier_elements) == 0:
            for i in range(0,3):
                try:
                    with allure.step(f"Attempt for driver refresh for campagins to load:- {i}"):
                        if i == 0:
                            driver.refresh()
                            time.sleep(5)
                            dashboard_loading(driver)
                            count_find_the_campaign = 1
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        elif i == 1:
                            logout(driver)
                            time.sleep(1.2)
                            clickec(driver, Login_Logout.link_login)
                            time.sleep(1.2)
                            login(driver, userid, password)
                            time.sleep(1.2)
                            count_find_the_campaign = 2
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        elif i == 2:
                            logout(driver)
                            time.sleep(1.2)
                            clickec(driver, Login_Logout.link_login)
                            time.sleep(1.2)
                            login(driver, userid, password)
                            driver.refresh()
                            time.sleep(5)
                            dashboard_loading(driver)
                            count_find_the_campaign = 3
                            click_on_side_bar_menu_compnents(driver, device, userid, password,Variable_MobileDevice_Xpath, classifier)
                        allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for driver refresh for campagins to load:- {i}",attachment_type=allure.attachment_type.PNG)
                        classifier_elements = driver.find_elements(classifier[0], classifier[1])
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                        if len(classifier_elements) != 0:
                            break
                except:
                    continue
        classifier_elements = driver.find_elements(classifier[0], classifier[1])
        for classifier_element in classifier_elements:
            action.move_to_element(classifier_element).perform()
            break
    except:
        pass
    finally:
        return count_find_the_campaign
def click_on_campaigns(driver,device,classifier,campaign,count_find_the_campaign,excelpath):
    try:
        assert clickec(driver, classifier)
        if count_find_the_campaign == 0:
            updatecomponentstatus("Side bar menu",str(campaign), "PASSED", f"{campaign} is found in {count_find_the_campaign} Attempt without Driver refresh.",excelpath)
        elif count_find_the_campaign == 1 or count_find_the_campaign == 2 or count_find_the_campaign == 3:
            updatecomponentstatus("Side bar menu",str(campaign), "FAILED", f"{campaign} is found in {count_find_the_campaign} Attempt after Driver refresh.",excelpath)
    except Exception as e:
        updatecomponentstatus("Side bar menu",str(campaign), "FAILED", f"Failed click on the {campaign} and check {campaign} is present in this {device} device",excelpath)
        if count_find_the_campaign == 1 or count_find_the_campaign == 2 or count_find_the_campaign == 3:
            updatecomponentstatus("Side bar menu", str(campaign), "FAILED",f"{campaign} is not found in {count_find_the_campaign} Attempt after Driver refresh.",excelpath)
        format_workbook(excelpath)
        raise e
    try:
        wait_for_loading_elements(driver)
    except:
        pass
    allure.attach(driver.get_screenshot_as_png(),name=f"Selected the device {str(device)} ==>> {str(campaign)} successfully",attachment_type=allure.attachment_type.PNG)

def click_on_side_bar_menu_compnents(driver,device,userid,password,Variable_MobileDevice_Xpath,classifier):
    classifier_element = "None"
    click_on_androidtestdata(driver)
    protestdata_runvalue = protestdata_module_controllers()
    litetestdata_runvalue = litetestdata_module_controllers()
    if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
        click_on_protestdata(driver)
    elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
        click_on_litetestdata(driver)
    try:
        MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
        # action = ActionChains(driver)
        # action.move_to_element(MobileDevice_element).perform()
        # simulate_screen_touch(100, 100)
    except:
        pass
    click_on_device(driver,userid,password,Variable_MobileDevice_Xpath,device)
    time.sleep(3.2)
    try:
        MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
        # action = ActionChains(driver)
        # action.move_to_element(MobileDevice_element).perform()
        # simulate_screen_touch(100, 100)
    except:
        pass
    search_campaigns(driver, classifier)
    time.sleep(2.2)
    assert uncheck_listOfcampaign(driver, side_menu_Components.campaignCheckBox)
    try:
        allure.attach(driver.get_screenshot_as_png(), name=f"Successfully unselected 'List of Campaign' checkbox",attachment_type=allure.attachment_type.PNG)
    except:
        pass
def search_campaigns(driver,classifier):
    action = ActionChains(driver)
    result = True
    try:
        elements = driver.find_elements(By.TAG_NAME, "a")
        for x in elements:
            if x.text == "Show More":
                result = True
    except Exception as e:
        pass
    max_run_time = 60
    start_time = time.time()
    while result == True:
        result = False
        try:
            classifier_element = None
            classifier_elements = driver.find_elements(classifier[0], classifier[1])
            try:
                for classifier_element in classifier_elements:
                    action.move_to_element(classifier_element).perform()
                    break
            except:
                pass
            if len(classifier_elements) != 0 and classifier_element.is_displayed():
                break
        except:
            pass
        time.sleep(0.01)
        try:
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.presence_of_element_located((side_menu_Components.element)))
        except Exception as e:
            pass
        elements = driver.find_elements(By.TAG_NAME, "a")
        for x in elements:
            if x.text == "Show More":
                try:
                    classifier_element = None
                    classifier_elements = driver.find_elements(classifier[0], classifier[1])
                    try:
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                    except:
                        pass
                    if len(classifier_elements) != 0 and classifier_element.is_displayed():
                        break
                except:
                    pass
                action.move_to_element(x).perform()
                x.click()
                result = True
                time.sleep(1)
                try:
                    with allure.step("classifers loading"):
                        wait_for_loading_elements(driver)
                except:
                    pass
                try:
                    classifier_element = None
                    classifier_elements = driver.find_elements(classifier[0], classifier[1])
                    try:
                        for classifier_element in classifier_elements:
                            action.move_to_element(classifier_element).perform()
                            break
                    except:
                        pass
                    if len(classifier_elements) != 0 and classifier_element.is_displayed():
                        break
                except:
                    pass
def wait_for_loading_elements(driver):
    timeout = 10
    # Define the loading element locators
    wait = WebDriverWait(driver, timeout)
    with allure.step("Waiting for application to load complete"):
        try:
            for by, locator in load_locators.loading_locators:
                wait.until(EC.invisibility_of_element_located((by, locator)))
        except Exception as e:
            allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for application to load completely",attachment_type=allure.attachment_type.PNG)
            print("An error occurred while waiting for loading elements:", str(e))
######################################################################################################################################
def remaining_test_minute_extraction(driver,statement):
    with allure.step(statement):
        allure.attach(driver.get_screenshot_as_png(), name=statement,attachment_type=allure.attachment_type.PNG)
        try:
            click(driver=driver, locators=Login_Logout.dashboard_id)
            remaing_test_min_locator = dashbord_more_info_components.remaing_test_min[:2]
            remaing_test_min_component_name = dashbord_more_info_components.remaing_test_min[2]
            remaining_test_minute_flag = wait_for_all_elements_presence(driver,locators=remaing_test_min_locator,timeout=10)
            try:
                remaining_test_minute_numeric_flag = wait_for_numeric_value(driver, remaing_test_min_locator, timeout = 10)
            except Exception as e:
                pass
            if remaining_test_minute_flag:
                remaining_test_minute_element = find_element(driver, locators=remaing_test_min_locator)
                remaining_test_minute_text = remaining_test_minute_element.text
                print(remaining_test_minute_text)
                allure.attach(driver.get_screenshot_as_png(), name=f"{remaining_test_minute_text}",attachment_type=allure.attachment_type.PNG)
                return remaining_test_minute_text
        except Exception as e:
            pass
############################################################################## Map view script #######################################################################################################
def Map_view_for_datetime_query(driver,excelpath,subtitle):

    driver.implicitly_wait(5)
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components_datetime_query(map_start_point, graph_start_point)
    # tests= ['Call Test', 'Failed Call', 'Ping Test', 'Web test', 'Download Test', 'Upload Test', 'HTTP DL', 'HTTP UL', 'TCPiperfDl', 'TCPiperfUl', 'UDPiperfDl', 'UDPiperfUl', 'Sent SMS', 'Received SMS', 'Failed SMS', 'Stream Test']
    Map_view(driver, tests, excelpath,subtitle,run_sub_modules=["Operator Comparison", "Interaction with Blob", "Distribution Graph"])
    driver.implicitly_wait(30)

def Map_view(driver, tests, excelpath, subtitle, run_sub_modules:Optional[List[str]] = None):
    if run_sub_modules is None:
        run_sub_modules = ["Operator Comparison", "Interaction with Blob", "Distribution Graph", "Floor Plan"]

    map_view_status_flag = None
    e_flag = "None"
    Notestdatafound_elements = "None"
    Title = f"{subtitle}-->MAP VIEW"
    try:
        # Notestdatafound_elements = driver.find_elements(*select_Map_View_Components.No_test_data_element)
        try:
            while WebDriverWait(driver,2).until(EC.visibility_of_element_located(select_Map_View_Components.No_test_data_element)):
                if WebDriverWait(driver,3).until(EC.invisibility_of_element_located(select_Map_View_Components.No_test_data_element)):
                    break
        except Exception as e:
            pass
        try:
            enable_of_element_untill_loaded(driver, select_Map_View_Components.Expand_Map_View[:2], 1)
            clickec(driver, select_Map_View_Components.Expand_Map_View)
            while WebDriverWait(driver, 1).until(EC.invisibility_of_element_located(close_button.closeFullTableView[:2])):
                clickec(driver, select_Map_View_Components.Expand_Map_View)
                if WebDriverWait(driver, 3).until(EC.visibility_of_element_located(close_button.closeFullTableView[:2])):
                    break
        except Exception as e:
            pass
            # Load pattern mapping from Excel file
        if WebDriverWait(driver, 3).until(EC.invisibility_of_element_located(select_Map_View_Components.No_test_data_element)):
            try:
                pattern_mapping_df = pd.read_excel(config.map_view_components_excelpath)
            except Exception as e:
                with allure.step(f"Check {config.map_view_components_excelpath}"):
                    print(f"Check {config.map_view_components_excelpath}")
                    assert False
            # Convert pattern mapping to dictionary
            pattern_mapping = pattern_mapping_df.set_index('Map view Components').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
            # Match patterns with tests
            txt = []
            if tests.__len__() == 0:
                statement  = f"Map-View  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
                with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for 'Map-View"):
                    # updatename(excelpath, statement)
                    updatecomponentstatus("MAP VIEW", "", "FAILED", f"Nothing marked in {str(config.test_data_path)}", excelpath)
                    e = Exception
                    raise e
            else:
                for test in tests:
                    test = test.strip()  # Remove leading and trailing spaces from test
                    for pattern, values in pattern_mapping.items():
                        if pattern.lower() == test.lower():
                            txt = values
                            break
                        else:
                            txt=[]
                    time.sleep(0.1)
                    try:
                        try:
                            listbox = WebDriverWait(driver,0.1).until(EC.visibility_of_element_located(select_Map_View_Components.map_menu_dropdown))
                            if listbox.is_displayed():
                                listbox_btn = WebDriverWait(driver, 1.2).until(EC.visibility_of_element_located(select_Map_View_Components.Test_Type_Dropdown))
                                # Click on the listbox to close it
                                listbox_btn.click()
                        except:
                            pass
                        Map_view_Search_Box_not_visible_do_page_up_(driver)
                        Map_View_Select_and_ReadData_(driver, select_Map_View_Components.Test_Type_Dropdown, select_Map_View_Components.nested_locators1, select_Map_View_Components.Call_Test_locator, txt, select_Map_View_Components.cluster_blobmap_locator, select_Map_View_Components.blobmap, select_Map_View_Components.map_element,floorplan.floormap ,test, select_Map_View_Components.Data_Table,Title,excelpath,test,run_sub_modules=run_sub_modules)
                    except Exception as e:
                        continue
            # click_closeButton(driver)
        # elif len(closeFullTableView_elements) == 0:
        elif WebDriverWait(driver, 3).until(EC.visibility_of_element_located(select_Map_View_Components.No_test_data_element)):
            statement = f"Failed to click on the expand for {Title}"
            with allure.step(statement):
                allure.attach(driver.get_screenshot_as_png(), name=f"Expand_Map_View_screenshot", attachment_type=allure.attachment_type.PNG)
                e = Exception
                raise e
    except Exception as e:
        Notestdatafound_elements = driver.find_elements(By.XPATH,"// h3[contains(text(), 'No test data found. Please try different date and ')]")
        closeFullTableView_elements = driver.find_elements(close_button.closeFullTableView[0],close_button.closeFullTableView[1])
        if len(closeFullTableView_elements) == 0:
            statement = f"Failed to click on the expand for {Title}"
            updatecomponentstatus(Title, "Expand_Map_View", "FAILED", statement, excelpath)
        elif e_flag == 1:
            print('select Map View Components fail')
        elif len(Notestdatafound_elements) != 0:
            statement = f"No test data found. Please try different date in Map View is present due to map didnt loaded"
            updatecomponentstatus(Title, "No test data found. Please try different date", "FAILED", statement, excelpath)
    finally:
        try:
            click_closeButton(driver)
        except Exception as e:
            pass

def click_closeButton(driver):
    try:
        Map_view_Search_Box_not_visible_do_page_up_(driver)
        clickec(driver,close_button.closeFullTableView)
    except Exception as e:
        raise e
def Map_view_Search_Box_not_visible_do_page_up_(driver):
    try:
        # Wait for the element to be visible
        driver.execute_script(f"window.scrollTo({0}, {0});")
    except:
        pass
                                 # driver, select_Map_View_Components.Test_Type_Dropdown, select_Map_View_Components.nested_locators1, select_Map_View_Components.Call_Test_locator, txt, select_Map_View_Components.cluster_blobmap_locator, select_Map_View_Components.blobmap, select_Map_View_Components.map_element, test, select_Map_View_Components.Data_Table, Title, excelpath, test, run_sub_modules
def Map_View_Select_and_ReadData_(driver, listbox_locator, nested_locators1, Call_Test_locator, option_text_list,cluster_blobmap_locator, blobmap_locator,floormap,map_element, elementname, table, Title,excelpath, test,run_sub_modules:List[str]):
    ListboxSelectstatus = "None"
    with allure.step(f"Map View Select '{elementname}' and Read Data"):
        try:

            l_flag = 0
            alert_text = None

            l_flag = 0
            if option_text_list.__len__() == 0:
                with allure.step(f"In input data from {str(config.map_view_components_excelpath)} for 'Map-View for '{test}' in header of Map view Components column value against the 2nd row of headers of Map view in {str(config.test_data_path)} is mismatch/empty"):
                    l_flag = 2
                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                    e = Exception
                    raise e
            elif ["Call Test", "Call Test"] != option_text_list and ['Call Test', 'Failed Calls'] != option_text_list:
                ListboxSelectstatus, alert_text = select_from_listbox_ECs(driver, listbox_locator, nested_locators1,option_text_list, Title, excelpath)
                l_flag = 1
            elif ["Call Test", "Call Test"] == option_text_list:
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Test_Type_Dropdown_for_call_Test, Title,excelpath)
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Call_Test_locator2, Title, excelpath)
                ListboxSelectstatus, alert_text = clickEC_for_listbox(driver, Call_Test_locator, Title, excelpath)
                l_flag = 1
            elif ['Call Test', 'Failed Calls'] == option_text_list:
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Test_Type_Dropdown_for_call_Test, Title,excelpath)
                clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Call_Test_locator2, Title, excelpath)
                ListboxSelectstatus, alert_text = clickEC_for_listbox(driver, Map_View_Select_and_ReadData.Failed_calls_locator,Title, excelpath)

                l_flag = 1

            # Handle operator comparison table and reduce redundant checks
            if alert_text is None and l_flag == 1:
                with allure.step("Operator Comparison"):
                    if "Operator Comparison" in run_sub_modules:
                        try:
                            while WebDriverWait(driver, 1).until(EC.invisibility_of_element_located(operator_comparison_table.operator_comparison_table)):
                                if WebDriverWait(driver, 2).until(EC.visibility_of_element_located(operator_comparison_table.operator_comparison_table)):
                                    break
                        except:
                           pass
                        try:
                            while WebDriverWait(driver, 2).until(EC.visibility_of_element_located(operator_comparison_table.operator_comparison_table)):
                                break
                        except:pass
                        try:
                            # Try to find and check the Webtest and comparison data elements in one go
                            elements_to_check = [
                                driver.find_elements(*operator_comparison_table.operator_comparison_web),
                                driver.find_elements(*operator_comparison_table.Operator_comparison_data),
                                driver.find_elements(*operator_comparison_table.operator_comparison_web_siblingtable)
                            ]

                            # Flatten the list of lists into one list
                            elements_to_check = [item for sublist in elements_to_check for item in sublist]

                            # Check if any element is displayed with non-empty text
                            if any(element.is_displayed() and element.text.strip() != "" for element in elements_to_check):
                                statement = f"Operator comparison table is present for {option_text_list[-1]}"
                                updatecomponentstatus(Title, str(test), "PASSED", statement, excelpath)
                            else:
                                statement = f"Operator comparison table is not present for {option_text_list[-1]}"
                                allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus(Title, str(test), "FAILED", statement, excelpath)
                                # Failupdatename(excelpath, statement)
                        except Exception as e:
                            pass  # Continue processing even if there's an exception
                    elif "Operator Comparison" not in run_sub_modules:
                        with allure.step("Not selected 'Operator Comparison'"):
                            pass

                # Handle blob map interaction
                with allure.step("Interaction with Blob"):
                    if "Interaction with Blob" in run_sub_modules:
                        try:
                            WebDriverWait(driver, 1).until(EC.visibility_of_element_located(blobmap_locator))
                        except:
                           pass
                        try:
                            cluster_blobmaps = driver.find_elements(*cluster_blobmap_locator)
                            if any(cluster_blobmap.is_displayed() for cluster_blobmap in cluster_blobmaps):
                                updatecomponentstatus(Title, str(test), "PASSED", "Blob Found", excelpath)
                            else:
                                blob_found_flag = interact_with_blobmap(driver, blobmap_locator, map_element, elementname)
                                if blob_found_flag == None or blob_found_flag == "Blob not found":
                                    updatecomponentstatus(Title, str(test), "FAILED", "Blob not found", excelpath)
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                                elif blob_found_flag == "Blob found":
                                    updatecomponentstatus(Title, str(test), "PASSED", "Blob Found", excelpath)
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot", attachment_type=allure.attachment_type.PNG)
                                else:
                                    updatecomponentstatus(Title, str(test), "WARNING",f"{blob_found_flag} should be handled", excelpath)
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                        except Exception as e:
                            updatecomponentstatus(Title, str(test), "WARNING", f"Should be handled {e}", excelpath)
                            pass
                    elif "Interaction with Blob" not in run_sub_modules:
                        with allure.step("Not selected 'Interaction with Blob'"):
                            pass

                with allure.step("Distribution Graph"):
                    if "Distribution Graph" in run_sub_modules:
                        try:
                            driver.execute_script(f"window.scrollTo({int(0)}, {int(200)})")
                            # time.sleep(3)
                            elements_to_check = [
                                driver.find_elements(*select_Map_View_Components.distribution_graph_other_test),
                                driver.find_elements(*select_Map_View_Components.distribution_graph_webtest),
                                driver.find_elements(*select_Map_View_Components.distribution_graph_smstest) ]
                            # Flatten the list of lists into one list
                            elements_to_check = [item for sublist in elements_to_check for item in sublist]

                            # Check if any element is displayed with non-empty text
                            if any(element.is_displayed()  for element in elements_to_check):
                                statement = f"Distribution graph is present for {option_text_list[-1]}"
                                updatecomponentstatus(Title, str(test), "PASSED", statement, excelpath)
                            else:
                                statement = f"Distribution graph is not present for {option_text_list[-1]}"
                                allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus(Title, str(test), "FAILED", statement, excelpath)
                        except Exception as e:
                            pass  # Continue processing even
                    elif "Distribution Graph" not in run_sub_modules:
                        with allure.step("Not selected 'Distribution Graph'"):
                            pass
                with allure.step("Floor Plan"):
                    if "Floor Plan" in run_sub_modules:
                        try:
                            floorplan_btn_element = driver.find_element(floorplan.floorplan_btn[0],floorplan.floorplan_btn[1])
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});",floorplan_btn_element)
                            clickec(driver,floorplan.floorplan_btn)
                            try:
                                floorplan_map = driver.find_elements(*floormap)
                                if any(floorplan_maps.is_displayed() for floorplan_maps in floorplan_map):
                                    updatecomponentstatus(Title, str(test), "PASSED", "Floor Plan found", excelpath)
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                                elif not any(floorplan_maps.is_displayed() for floorplan_maps in floorplan_map):
                                    updatecomponentstatus(Title, str(test), "FAILED", "Floor Plan found",excelpath)
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
                            except Exception as e:
                                updatecomponentstatus(Title, str(test), "WARNING", f"Should be handled {e}",excelpath)
                                pass
                        except Exception as e:
                            pass  # Continue processing even
                    elif "Floor Plan" not in run_sub_modules:
                        with allure.step("Not selected 'Floor Plan'"):
                            pass
            else:
                handle_alert_or_failure_case(driver, alert_text, l_flag, excelpath, elementname, Title,option_text_list, test)

        except Exception as e:
            print("Map View Select and Read Data fail")
            raise e

def handle_alert_or_failure_case(driver, alert_text, l_flag, excelpath, elementname, Title, option_text_list, test):
    if l_flag == 0:
        statement = f"Unable to locate the element/No such element found for {option_text_list} from listbox"
    elif l_flag == 2:
        statement = f"In input data from {config.map_view_components_excelpath} for 'Map-View for '{test}' is mismatch/empty"
    elif alert_text is not None and l_flag == 1:
        statement = f"Alert Found: '{alert_text}' for Map View to select {elementname}"
    else:
        statement = "Unknown error occurred"

    allure.attach(driver.get_screenshot_as_png(), name=f"{elementname}_screenshot",attachment_type=allure.attachment_type.PNG)
    updatecomponentstatus(Title, elementname, "FAILED", statement, excelpath)
    # Failupdatename(excelpath, statement)
    raise Exception(statement)


###################################################### Date and Time Requirement functions are below #################################################################################################
def click_load_more(driver,load_more_button_xpath,time):
    while True:
        try:
            enable_of_element_untill_loaded(driver, load_more_button_xpath[:2], 1)
            load_more_button = WebDriverWait(driver, time).until(EC.visibility_of_element_located(load_more_button_xpath[:2]))
            clickec(driver,load_more_button_xpath)
            time.sleep(20)
        except:
            break

    try:
        while WebDriverWait(driver, 10).until(EC.presence_of_element_located(load_more_button_xpath[:2])):
            try:
                enable_of_element_untill_loaded(driver, load_more_button_xpath[:2], 1)
                load_more_button = WebDriverWait(driver, time).until(EC.presence_of_element_located(load_more_button_xpath[:2]))
                load_more_button.click()
            except Exception as e:
                break
    except Exception as e:
        pass
###########################################################################################################################################################################
def navigate_to_date(driver, start_date, end_date):
    # Extract year, month, and day from start_date
    start_year = start_date.year
    start_month = start_date.month
    start_day = start_date.day

    # Extract year, month, and day from end_date
    end_year = end_date.year
    end_month = end_date.month
    end_day = end_date.day

    start_months_difference = current_difference_in_months(start_year, start_month)
    end_months_difference = userdefined_difference_in_months(end_year,end_month,start_year, start_month)
    # if start_year < current_date.year or start_month < current_date.month:
    if start_months_difference != 0:
        for i in range(0,start_months_difference):

            click(driver,date_time.left_calender_previous_btn)
    left_calender_date = (By.XPATH,f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar second left']//div[@class='calendar-date']//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{start_day}']","left_calender_date")
    right_calender_date = (By.XPATH,f"//div[contains(@class, 'daterangepicker') and contains(@style, 'display: block;')]//div[@class='calendar first right']//div//td[contains(@class, 'available') and not(contains(@class, 'available off')) and text()='{end_day}']","right_calender_date")

    click(driver,left_calender_date)
    # if end_year < current_date.year or end_month < current_date.month:
    if end_months_difference != 0:
        for i in range(0,end_months_difference):

            click(driver,date_time.right_calender_forward_btn)

    click(driver, right_calender_date)
    click(driver, date_time.datetime_apply_btn)

def current_difference_in_months(year, month):
    from datetime import datetime
    # Get the current year and month
    current_year = datetime.now().year
    current_month = datetime.now().month
    # Calculate the difference in months
    difference = (current_year - year) * 12 + (current_month - month)
    return difference

def userdefined_difference_in_months(year1,month1,year2, month2):
    # Calculate the difference in months
    difference = (year1 - year2) * 12 + (month1 - month2)
    return difference
#################################################################################################################################################
def extract_table_column_data(driver,excelpath,Title):
    try:
        # Wait for the table to be present
        WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, "//div[@class='div-table-content-wrapper']")))
    except Exception as e:
        print("Error waiting for table to be present:", e)
    individual_pop_loadcamp_element = driver.find_elements(*individual_pop_table.loaderCamp[:2])
    data = []
    for i in range(len(individual_pop_loadcamp_element)):
        i += 1
        individual_pop_table_loaderCamp = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/abbr/a", "individual_pop_table_loaderCamp")
        test_name_xpath = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[1]")
        Device_name_xpath = (By.XPATH, f"//tr[{i}]//*[@id='loaderCamp']/following-sibling::td[2]")
        individual_pop_table_loaderCamp_element = driver.find_element(*individual_pop_table_loaderCamp[:2])
        individual_pop_table_loaderCamp_name = individual_pop_table_loaderCamp_element.text
        test_name_element = driver.find_element(*test_name_xpath)
        test_name = test_name_element.text
        Device_name_element = driver.find_element(*Device_name_xpath)
        Device_name = Device_name_element.text
        data.append({
            'Operator Name': individual_pop_table_loaderCamp_name, 'Test Name': test_name, 'Device': Device_name # "Device" in the third column (index 2)
        })
        updatecomponentstatus(Title, f"'Operator Name', 'Test Name', 'Device'", "",f"'Operator Name': {individual_pop_table_loaderCamp_name}, 'Test Name': {test_name}, 'Device': {Device_name}",excelpath)
#######################################################################################################################################################################################################################
def date_and_time_main_function(driver,excelpath):
    Title = "Date and Time"
    datetime_runvalue = date_and_time_module_controllers()
    if "Yes".lower() == datetime_runvalue[-1].strip().lower():
        try:
            try:
                clickec(driver, date_time.date_and_time_click_button)
            except Exception as e:
                print("Could not able to click")
            df = pd.read_excel(config.test_data_path, sheet_name='date_time')
            timerequired =  False
            # Loop through each row in the DataFrame
            for index, row in df.iterrows():
                select_hours = row['Select_hours']
                execute_flag = row['Execute']
                if isinstance(execute_flag, str) and execute_flag.lower() == 'yes':
                    if select_hours.lower() == 'custom date':
                        # Handle custom date range
                        start_date = row['Start Date']
                        end_date = row['End Date']

                        # Click on the button to open the date picker
                        click(driver, date_time.date_picker_button_xpath)

                        # code to select custom date range
                        navigate_to_date(driver, start_date,end_date)
                        click(driver, date_time.datetime_apply_btn)
                        timerequired = True
                    else:
                        try:
                            radio_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((date_time.radio_button_xpath[0],date_time.radio_button_xpath[1].format(str(select_hours).lower()))))
                            radio_button.click()
                        except Exception as e:
                            print(f"Failed to select {select_hours}: {str(e)}")
            Page_Down(driver)
            clickec(driver,date_time.expand_table_button)
            Page_up(driver)
            try:
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(date_time.load_more_button_xpath))
                click_load_more(driver, date_time.load_more_button_xpath, 20)
            except Exception as e:
                for i in range(0, 3):
                    time.sleep(1)
                    clickec(driver=driver, locators=remote_test.table_view_refresh)
            try:
                # with allure.step("Date and Time table data extract"):
                 extract_table_column_data(driver,excelpath,Title)
                 statement = f"Successfully extracted the table data "
                 with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"Date and Time",attachment_type=allure.attachment_type.PNG)
            except:
                statement = f"Failed to  extracted the table data "
                with allure.step(statement):
                    allure.attach(driver.get_screenshot_as_png(), name=f"Date and Time",attachment_type=allure.attachment_type.PNG)
            clickec(driver, close_button.closeFullTableView)
            Map_view_for_datetime_query(driver, excelpath,subtitle = "Date and Time")

        except:
            pass

        finally:
             status = readcomponentstatus_(status_word = "WARNING|FAILED", path = excelpath, Titlename = f"{Title}-->MAP VIEW",condition="eq")
             if status== "FAILED":
                 updatehighmodulestatus(Title,status,comments = "Date and Time query failed",path=excelpath)
             elif status == "PASSED":
                 updatehighmodulestatus(Title, status, comments="Date and Time query Passed", path=excelpath)

    elif "No".lower() == datetime_runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute", excelpath)
            updatehighmodulestatus(Title, "SKIPPED", "You have selected No for execute", excelpath)
            pass

################################################################ Device - Custom Query ####################################################################################################################
def device_custom_query(driver, campaigns_datas, userid, password,excelpath):
    Title = "DEVICE CUSTOM QUERY"
    device_custom_query_runvalue = device_custom_query_module_controllers()

    if "Yes".lower() == device_custom_query_runvalue[-1].strip().lower():
        try:
            side_menu_custom_query(driver, campaigns_datas, userid, password)
            selecting_date_or_hours_for_device(driver,Title,excelpath)
            load_more_extract_data(driver, excelpath,Title)
            Map_view_components_device_customquery(driver,excelpath,subtitle = "Device Custom Query")
        except:
            pass
        finally:
             status = readcomponentstatus_(status_word="WARNING|FAILED", path=excelpath, Titlename=f"{Title}-->MAP VIEW",condition="eq")
             if status == "FAILED":
                 updatehighmodulestatus(Title, status, comments="Device - Custom Query failed", path=excelpath)
             elif status == "PASSED":
                 updatehighmodulestatus(Title, status, comments="Device - Custom Query Passed", path=excelpath)
    elif "No".lower() == device_custom_query_runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute",excelpath)
            updatehighmodulestatus(Title, "SKIPPED", "You have selected No for execute", excelpath)
            pass

################################################### side bar menu --> Android --> Protest --> Device ####################################################
def side_menu_custom_query(driver, campaigns_datas, userid, password):
    try:
        device_list = [campaigns_datas[i][0] for i in range(len(campaigns_datas))]
        device_list = list(set(device_list))
        for device in device_list[:1]:
            Variable_MobileDevice_Xpath = (By.XPATH, f"//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
            protestdata_runvalue = protestdata_module_controllers()
            litetestdata_runvalue = litetestdata_module_controllers()
            if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
                Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='Pro TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
            elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
                Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='LITE TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
            click_on_android_protest_device(driver, device, userid, password, Variable_MobileDevice_Xpath)
    except Exception as e:
        with allure.step("Failed in side bar menu"):
            allure.attach(driver.get_screenshot_as_png(), name=f"Failed in side bar menu",attachment_type=allure.attachment_type.PNG)
            pytest.fail(str(e))
def click_on_android_protest_device(driver,device,userid,password,Variable_MobileDevice_Xpath):
    click_on_androidtestdata(driver)
    protestdata_runvalue = protestdata_module_controllers()
    litetestdata_runvalue = litetestdata_module_controllers()
    if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
        click_on_protestdata(driver)
    elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
        click_on_litetestdata(driver)
    try:
        MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
    except:
     pass
    click_on_device(driver,userid,password,Variable_MobileDevice_Xpath,device)
    # time.sleep(3.2)
    assert uncheck_listOfcampaign(driver, side_menu_Components_custom_query.campaignCheckBox)
    try:
        allure.attach(driver.get_screenshot_as_png(), name=f"Successfully unselected 'List of Campaign' checkbox",attachment_type=allure.attachment_type.PNG)
    except:
        pass
def click_on_androidtestdata(driver):
    for i in range(0,2):
        try:
           with allure.step(f"Attempt for androidtestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                # time.sleep(1.2)
                androidtestdata_element = driver.find_elements(side_menu_Components_custom_query.androidtestdata[0],side_menu_Components_custom_query.androidtestdata[1])
                if len(androidtestdata_element)!=0:
                    assert clickec(driver, side_menu_Components_custom_query.androidtestdata)
                    # time.sleep(1.2)
                    protestdata_runvalue = protestdata_module_controllers()
                    if protestdata_runvalue[0].lower() == 'Yes'.lower():
                        protestdata_element = driver.find_elements(side_menu_Components_custom_query.protestdata[0],side_menu_Components_custom_query.protestdata[1])
                        if len(protestdata_element) != 0:
                            break
                    elif protestdata_runvalue[0].lower() == 'No'.lower():
                        litetestdata_element = driver.find_elements(side_menu_Components_custom_query.litetestdata[0],side_menu_Components_custom_query.litetestdata[1])
                        if len(litetestdata_element) != 0:
                            break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for androidtestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
def click_on_protestdata(driver):
    for i in range(0,2):
        try:
            with allure.step(f"Attempt for protestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                # time.sleep(1.2)
                protestdata_element = driver.find_elements(side_menu_Components_custom_query.protestdata[0],side_menu_Components_custom_query.protestdata[1])
                if len(protestdata_element) != 0:
                    assert clickec(driver, side_menu_Components_custom_query.protestdata)
                    # time.sleep(1.2)
                    device_element = driver.find_elements(*side_menu_Components_custom_query.device_element)
                    if len(device_element) != 0:
                        break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for protestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue

def click_on_litetestdata(driver):
    for i in range(0,2):
        try:
            with allure.step(f"Attempt for litetestdata :- {i}"):
                if i > 0:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                # time.sleep(1.2)
                litetestdata_element = driver.find_elements(side_menu_Components_custom_query.litetestdata[0],side_menu_Components_custom_query.litetestdata[1])
                if len(litetestdata_element) != 0:
                    assert clickec(driver, side_menu_Components_custom_query.litetestdata)
                    # time.sleep(1.2)
                    device_element = driver.find_elements(*side_menu_Components_custom_query.device_element)
                    if len(device_element) != 0:
                        break
                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for litetestdata :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
def click_on_device(driver,userid,password,Variable_MobileDevice_Xpath,device):
    protestdata_runvalue = protestdata_module_controllers()
    litetestdata_runvalue = litetestdata_module_controllers()
    for i in range(0,3):
        try:
            with allure.step(f"Attempt for device :- {i}"):
                if i == 1:
                   driver.refresh()
                   dashboard_loading(driver)
                   click_on_androidtestdata(driver)
                   if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
                       click_on_protestdata(driver)
                   elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
                       click_on_litetestdata(driver)
                elif i == 2:
                   logout(driver)
                   # time.sleep(1.2)
                   clickec(driver, Login_Logout.link_login)
                   # time.sleep(1.2)
                   login(driver, userid, password)
                   # time.sleep(1.2)
                   click_on_androidtestdata(driver)
                   if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
                       click_on_protestdata(driver)
                   elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
                       click_on_litetestdata(driver)
                # time.sleep(1.2)
                device_elements = driver.find_elements(By.XPATH, f"//a[normalize-space()='{str(device)}']")
                if len(device_elements) != 0:
                    assert clickec(driver, Variable_MobileDevice_Xpath)

                allure.attach(driver.get_screenshot_as_png(),name=f"Attempt for device :- {i}",attachment_type=allure.attachment_type.PNG)
        except:
            continue
########################################################################## Selection of time period #####################################################################################################################################
def selecting_date_or_hours_for_device(driver,Title,excelpath):
    with allure.step("Selecting date or hours for device custom query"):
        yes_flag = False
        selected_flag = False
        selected_value = None
        try:
            # Page_up(driver)
            df = pd.read_excel(config.test_data_path, sheet_name='date_time')
            # Loop through each row in the DataFrame

            for index, row in df.iterrows():
                select_hours = row['Select_hours']
                execute_flag = row['Execute']
                if isinstance(execute_flag, str) and execute_flag.lower() == 'yes':
                    yes_flag = True
                    if select_hours.lower() == 'custom date':
                        # Handle custom date range
                        start_date = row['Start Date']
                        end_date = row['End Date']
                        # Click on the button to open the date picker
                        click(driver, custom_query_device.custom_date)
                        # Code to select custom date range
                        navigate_to_date_for_device_custom(driver, start_date, end_date)
                        clickec(driver, custom_query_device.apply_btn)
                        updatecomponentstatus(Title=Title, componentname=f"{select_hours}=={start_date}/{end_date}", status="PASSED",comments="Successfully Selected",path=excelpath)
                        selected_flag = True
                        selected_value = f"{select_hours}=={start_date}/{end_date}"

                    else:
                        clickec(driver, custom_query_device.drop_down_device)
                        option_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((custom_query_device.drop_down_value[0],custom_query_device.drop_down_value[1].format(str(select_hours).lower()))))
                        option_element.click()
                        updatecomponentstatus(Title=Title, componentname=f"{select_hours}",status="PASSED", comments="Successfully Selected", path=excelpath)
                        selected_flag = True
                        selected_value = f"{select_hours}"

            if yes_flag == False:
                updatecomponentstatus(Title=Title, componentname=f"{selected_value}", status="FAILED",comments="Select anyone option in date/time by providing 'Yes' in '{config.test_data_path}' sheet:-'date_time'", path=excelpath)

            elif yes_flag == True and selected_flag == False:
                updatecomponentstatus(Title=Title, componentname=f"{selected_value}", status="FAILED", comments="Failed to select", path=excelpath)
            allure.attach(driver.get_screenshot_as_png(), name=f"Selecting date or hours for device custom query screenshot",attachment_type=allure.attachment_type.PNG)
        except Exception as e:
            if yes_flag == True and selected_flag == False:
                updatecomponentstatus(Title=Title, componentname=f"{selected_value}", status="FAILED", comments="Failed to select", path=excelpath)
            allure.attach(driver.get_screenshot_as_png(), name=f"Selecting date or hours for device custom query screenshot",attachment_type=allure.attachment_type.PNG)
            raise e
def navigate_to_date_for_device_custom(driver, start_date, end_date):
    # Extract year, month, and day from start_date
    global current_year
    start_year = start_date.year
    start_month = start_date.month
    start_day = start_date.day
    # Assuming start_month is the month number, for example, 5 for May
    start_month_number = start_month
    # Convert the month number to its corresponding name
    start_month_name = calendar.month_name[start_month_number][:3]
    # Loop until the start month and year are visible
    while True:
            try:
                try:
                    # Get the currently displayed month and year

                    current_month_year_element = driver.find_element(*custom_query_device.current_month_year_xpath)
                    # Get the currently displayed month and year
                    current_month_year = current_month_year_element.text
                    # Parse the displayed month and year
                    start_month_name, current_year = current_month_year.split()
                except Exception as e:
                    print(Exception)
                    pass
                start_month_name = list(calendar.month_abbr).index(start_month_name)
                # Check if the start month and year are visible
                if start_month_name == start_month and int(current_year) == start_year:
                    break
            except Exception as e:
                pass
            # Find and click the navigation button to go to the previous or next month
            if int(current_year) > start_year or (int(current_year) == start_year and start_month_name > start_month):
                # Click on the "<" button to go to the previous month
                navigation_button_xpath = custom_query_device.navigation_button_prevavailable_xpath
            else:
                # Click on the ">" button to go to the next month
                navigation_button_xpath = custom_query_device.navigation_button_nextavailable_xpath
            navigation_button = driver.find_element(*navigation_button_xpath)
            navigation_button.click()
    # Find and click the element corresponding to the start day
    start_day_element = driver.find_element(custom_query_device.start_day_element_xpath[0],custom_query_device.start_day_element_xpath[1].format(start_day))
    start_day_element.click()
    # # Extract year, month, and day from end_date
    end_year = end_date.year
    end_month = end_date.month
    end_day = end_date.day
    end_month_number = end_month
    end_month_name = calendar.month_name[end_month_number][:3]
    # Loop until the end month and year are visible
    while True:
        try:
            try:
                # Get the currently displayed month and year
                current_month_year_element = driver.find_element(*custom_query_device.current_month_year_xpath1)
                current_month_year = current_month_year_element.text
                # Parse the displayed month and year
                end_month_name, current_year = current_month_year.split()
            except Exception as e:
                pass
            end_month_name = list(calendar.month_abbr).index(end_month_name)
            # Check if the end month and year are visible
            if end_month_name == end_month and int(current_year) == end_year:
                break
        except Exception as e:
            pass
        # Find and click the navigation button to go to the previous or next month
        if int(current_year) > end_year or (int(current_year) == end_year and end_month_name > end_month):
            navigation_button_xpath1 = custom_query_device.navigation_button_prevavailable_xpath1
        else:
            navigation_button_xpath1 = custom_query_device.navigation_button_nextavailable_xpath1
        navigation_button = driver.find_element(*navigation_button_xpath1)
        navigation_button.click()
    # Find and click the element corresponding to the end day
    end_day_element = driver.find_element(custom_query_device.end_day_element_xpath[0],custom_query_device.end_day_element_xpath[1].format(end_day))
    end_day_element.click()

########################################################################################################################################################################################################################################
def load_more_extract_data(driver,excelpath,Title):
    try:
        element = driver.find_element(*select_Map_View_Components.expand_table_button[:2])
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    except:
      pass
    clickec(driver, select_Map_View_Components.expand_table_button)
    Page_up(driver)
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located(date_time.load_more_button_xpath))
        click_load_more(driver, date_time.load_more_button_xpath, 20)
    except Exception as e:
        for i in range(0, 3):
            time.sleep(1)
            clickec(driver=driver, locators=remote_test.table_view_refresh)
    extract_table_column_data(driver,excelpath,Title)
    clickec(driver, close_button.closeFullTableView)

############################################################## Verification in Map View ##############################################################################################################
def Map_view_components_device_customquery(driver,excelpath,subtitle):
    driver.implicitly_wait(5)
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components_datetime_query(map_start_point, graph_start_point)
    Map_view(driver, tests, excelpath, subtitle,run_sub_modules = ["Operator Comparison", "Interaction with Blob", "Distribution Graph"])
    driver.implicitly_wait(30)

#################################################### Export_Dashboard ######################################################################################################
def Export_Dashboard(driver,excelpath, campaign, downloadfilespath):
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, export_start_point, load_start_point)
    Exports_view(driver, tests, excelpath, downloadfilespath)

def Exports_view(driver,tests,excelpath,downloadfilespath):
    downloadfilespath = specifying_download_path(driver, downloadfilespath, "EXPORTS")
    Title = "Exports"
    exports_runvalue = exports_module_controllers()
    result_status = queue.Queue()
    if "Yes".lower() == exports_runvalue[-1].strip().lower():

        # Fetch components based on the campaign/classifier "T001","T002" etc
        if tests == []:
            List_of_options_txts=[["Combined Export"], ["Survey Test Export"], ["Export TableSummary"],["Export As PDF"],["Combined Binary Export"],["Hand OverExport"]]
        else:
            input_list = tests
            List_of_options_txts = [[item] for item in input_list]
        try:
            # ReportDownlaodName(excelpath, "*************    Download-Reports  --- Starts  from here  *************")
            for List_of_options_txt in List_of_options_txts:
                # traverse in the string
                exportname = ""
                for ele in List_of_options_txt:
                    exportname += ele
                try:
                    try:
                        if WebDriverWait(driver, 1).until(EC.visibility_of_element_located(List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown)).is_displayed():
                            listbox_btn = WebDriverWait(driver, 1).until(EC.visibility_of_element_located(List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown))
                            # Click on the listbox to close it
                            listbox_btn.click()
                    except:
                        pass
                    time.sleep(1.2)
                    List_Of_Campaigns_components(driver, List_of_options_txt, List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options, exportname, excelpath,result_status,downloadfilespath)
                except Exception as e:
                    continue
        except Exception as e:
            print('Exports_fail')
        try:
            with allure.step("updating_export_result_to_excel"):
                updating_export_result_to_excel(result_status,excelpath)
        except Exception as e:
            with allure.step("failed:- updating_export_result_to_excel"):
                pass

    elif "No".lower() == exports_runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus(Title, "Not to execute", "SKIPPED", "You have selected No for execute",excelpath)
            pass
def updating_export_result_to_excel(result_status,excelpath):
    dataframe_status =[]
    combined_status_df="None"
    while not result_status.empty():
        updatecomponentstatus2 = result_status.get()
        df_status = pd.DataFrame(updatecomponentstatus2)
        dataframe_status.append(df_status)
    if len(dataframe_status) !=0:
        with allure.step("result_status of export updating to excel"):
            combined_status_df = pd.concat(dataframe_status, ignore_index=True)
    workbook = openpyxl.load_workbook(excelpath)
    worksheet_componentstatus = workbook["COMPONENTSTATUS"]

    if len(dataframe_status) !=0:
        update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
    # while not result_data.empty():
    #     with allure.step("result_data of export updating to excel"):
    #         data = result_data.get()
    #         try:
    #             df_data = pd.DataFrame(data)
    #         except Exception as e:
    #             # Check if the values associated with keys are single scalars, and convert them to lists if needed
    #             for key, value in data.items():
    #                 if not isinstance(value, list):
    #                     data[key] = [value]
    #             df_data = pd.DataFrame(data)
    #         # List of columns to be moved to the front
    #         List_of_options_txts = [["Combined Export"], ["Survey Test Export"], ["Export TableSummary"], ["Export As PDF"],["Combined Binary Export"], ["Hand OverExport"]]
    #         # Find the first sublist item that is a key in df_data
    #         first_match = next((item[0] for item in List_of_options_txts if item[0] in df_data.columns), None)
    #         if first_match:
    #             # Reorder the columns in df_data with the first match at the beginning
    #             df_data = df_data[[first_match] + [col for col in df_data.columns if col != first_match]]
    #         updating_data_of_dataframe_for_excel(worksheet=worksheet_data_extract, df_data=df_data)
    workbook.save(excelpath)
    workbook.close()
def updating_data_of_dataframe_for_excel(worksheet,df_data):
    # Convert the DataFrame to a list of rows
    data = list(dataframe_to_rows(df_data, index=False, header=True))
    # Find the last row in the existing data and calculate the next available row for appending
    last_row = worksheet.max_row
    next_available_row = last_row + 1
    # Define a fill color for the first column (1st column value)
    first_column_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow color
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    # Iterate through rows and columns to append data to the worksheet
    for row_idx, row_data in enumerate(data, 1):  # Start from row 1
        for col_idx, cell_value in enumerate(row_data, 1):  # Start from column 1
            cell = worksheet.cell(row=next_available_row + row_idx, column=col_idx, value=cell_value)
            # Apply the color fill to the first column (assuming it's the first column)
            if col_idx == 1:
                cell.fill = first_column_fill
            # Apply the color fill to the header row (assuming it's the first row)
            if row_idx == 1:
                cell.fill = header_fill
def List_Of_Campaigns_components_Search_Box_not_visible_do_page_up_(driver):
    try:
        # Wait for the element to be visible
        wait = WebDriverWait(driver, 10)
        Search_Element = wait.until(EC.visibility_of_element_located(List_Of_Campaigns_components_Search_Box_not_visible_do_page_up.Search_Element))
        Search_Element.click()
    except:
        Page_up(driver)

def List_Of_Campaigns_components(driver,List_of_options_txt, List_Of_Campaigns_Export_Dropdown, List_Of_Campaigns_Export_Dropdown_Options, exportname, excelpath,result_status,downloadfilespath):
    Title = "Exports"
    with allure.step(f"List Of Campaigns Export '{exportname}' component"):
        try:
            Title = "Exports"
            # Navigate to the "Export As PDF" option
            if List_of_options_txt == ["Export As PDF"]:
                try:
                    # Store the original window handle
                    original_window_handle = driver.current_window_handle
                    time.sleep(1.2)
                    select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dropdown_Options, List_of_options_txt,Title, excelpath)
                    time.sleep(1.2)
                    # Switch to the new tab
                    driver.switch_to.window(driver.window_handles[1])
                    # Take a screenshot and attach it to the Allure report
                    with allure.step("Export As PDF Screenshot"):
                        allure.attach(driver.get_screenshot_as_png(), name="screenshot",attachment_type=allure.attachment_type.PNG)
                        updatecomponentstatus2= status(Title, List_of_options_txt.__str__(), "PASSED",f"Passed step :- Screenshot of Export As PDF is taken'")
                        result_status.put(updatecomponentstatus2)
                    # Close the second window
                    driver.close()
                    # Switch back to the original window
                    driver.switch_to.window(original_window_handle)
                except Exception as e:
                    with allure.step(f"Step Failed :- Screenshot of Export As PDF is not taken due PDF is not present"):
                        updatecomponentstatus2= status(Title, List_of_options_txt.__str__(), "FAILED",f"Step Failed :- Screenshot of Export As PDF is not taken due PDF is not present'")
                        result_status.put(updatecomponentstatus2)
                        raise e
            elif List_of_options_txt != ["Export As PDF"]:
                time.sleep(1.2)
                flag, alert_text = select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dropdown_Options, List_of_options_txt,Title, excelpath)
                time.sleep(4)
                if flag == 0 and alert_text == None:
                    readCSVSheet( Title, exportname,result_status,downloadfilespath)
                elif flag == 0 and alert_text != None:

                    e = Exception
                    with allure.step(f"Failed Step :- Alert Found is '{alert_text}' for List Of Campaigns Export '{exportname}' component"):
                        updatecomponentstatus2 = status(Title,exportname, "FAILED", f"failed step :- Alert Found :-'{alert_text}'")
                        result_status.put(updatecomponentstatus2)
                        raise e
        except Exception as e:
            raise e

######################################################## Map view(NQC-operator comparison V/S PDF Export) ###########################################################################################################
def Mapview_Operator_comparison_vs_PDF(driver,campaign,excelpath,downloadpdfpath,settingvalue,excelpath_for_storedata,data_match_sheet,data_not_match_sheet,remote_test_campaign,device):
    downloadpdfpath = specifying_download_path(driver, downloadpdfpath, "PDF")
    result_status = queue.Queue()
    data_difference = queue.Queue()
    data_same = queue.Queue()
    Title = "Map view(NQC-operator comparison V/S PDF Export)"
    remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
    tests = fetch_components(campaign, PDF_Export_index_start_point, END_index)
    tests_no = fetch_components_for_no_yes(campaign, PDF_Export_index_start_point, END_index)
    Operator_vs_pdf_runvalue = operatorcomparison_vs_pdf_module_controllers()
    if "Yes".lower() == Operator_vs_pdf_runvalue[-1].strip().lower():
        driver.implicitly_wait(5)
        pdf_comparsion(driver, tests, Title, result_status, tests_no, downloadpdfpath, data_difference, data_same, excelpath=excelpath,settingvalue=settingvalue,excelpath_for_storedata=excelpath_for_storedata,remote_test_campaign=remote_test_campaign,campaign=campaign,device=device)
        driver.implicitly_wait(30)
    elif "No".lower() == Operator_vs_pdf_runvalue[-1].strip().lower():
        statement = "You have selected Not to execute"
        with allure.step(statement):
            updatecomponentstatus2 = status(Title, "Not to execute", "SKIPPED", "You have selected No for execute")
            result_status.put(updatecomponentstatus2)
            pass
    try:
        with allure.step("update_result_of_pdf"):
            update_result_of_pdf(result_status, data_difference, data_same, excelpath,data_match_sheet,data_not_match_sheet)
    except Exception as e:
        with allure.step(f"failed step:- update_result_of_pdf {str(e)}"):
            pass

def pdf_comparsion(driver,tests,Title,result_status,tests_no,downloadpdfpath,data_difference,data_same,excelpath,settingvalue,excelpath_for_storedata,remote_test_campaign,campaign,device):
    enabled_checkboxes = None
    disabled_checkboxes = None
    checkbox_option_text_list = None
    pdf_files  = None
    try:
        List_of_options_txts = [["Export"],["Export As PDF"]]
        # Load pattern mapping from Excel file
        try:
            pattern_mapping_df = pd.read_excel(config.pdf_export_excel_path,sheet_name="pdf_components")
        except Exception as e:
            with allure.step(f"Check {config.map_view_components_excelpath}"):
                print(f"Check {config.map_view_components_excelpath}")
                assert False
        for List_of_options_txt in List_of_options_txts:
            flag, alert_text = select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
        # Convert pattern mapping to dictionary
        pattern_mapping = pattern_mapping_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
        # Match patterns with tests
        txt = []
        if tests.__len__() == 0:
            statement = f"{Title}  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
            with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for {Title}"):
                updatecomponentstatus2 = status(Title, "", "FAILED",f"Nothing marked in {str(config.test_data_path)}")
                result_status.put(updatecomponentstatus2)
                e = Exception
                raise e
        else:
            enabled_txt=[]
            txts = []
            for test in tests:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        txts = values
                        enabled_txt.append(txts)
                        break
                    else:
                        txts = []
            disabled_txt = []
            disabledtxts = []
            for test in tests_no:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        disabledtxts = values
                        disabled_txt.append(disabledtxts)
                        break
                    else:
                        disabledtxts = []
            s_flag = 0
            try:
                driver.switch_to.window(driver.window_handles[1])
                s_flag = 1
            except Exception as e:
                pass
            try:
                pdf_listbox_locator = (By.XPATH, "//div[@id='checkboxes']//input[@type='checkbox']")
                WebDriverWait(driver, 2).until(EC.visibility_of_element_located(pdf_listbox_locator))
            except Exception as e:
                try:
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    List_of_options_txts = [["Export"], ["Export As PDF"]]
                    allure.attach(driver.get_screenshot_as_png(), name="PDF Data Extraction",attachment_type=allure.attachment_type.PNG)
                    for List_of_options_txt in List_of_options_txts:
                        try:
                            select_from_listbox_ECs(driver,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
                        except:
                            pass
                    time.sleep(2)
                    driver.switch_to.window(driver.window_handles[1])
                except Exception as e:
                    pass
            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
            try:

                start_time = time.time()
                # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                max_run_time = 60
                if len(pdf_export_checkbox) == 0:
                    with allure.step("Waiting for pdf page to load"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for pdf page to load",attachment_type=allure.attachment_type.PNG)
                        while time.time() - start_time < max_run_time:
                            try:
                                driver.switch_to.window(driver.window_handles[1])
                            except Exception as e:
                                pass
                            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
                            # Check if the condition is met
                            if len(pdf_export_checkbox) != 0:
                                break
            except:
                pass

            time.sleep(5)
            enabled_checkboxes, disabled_checkboxes = check_selected_and_finding_enable_and_disabled_checkboxes_(driver,pdf_view.parent_checkbox_pdf)
            with concurrent.futures.ThreadPoolExecutor() as executor:
                for enabled_checkboxs in enabled_txt:
                    executor.submit(process_enabled_checkbox, Title, result_status, enabled_checkboxes,enabled_checkboxs, disabled_checkboxes)
            with concurrent.futures.ThreadPoolExecutor() as executor:
                for disabled_checkboxs in disabled_txt:
                    executor.submit(process_disabled_checkbox, Title, result_status, disabled_checkboxs,enabled_checkboxes, disabled_checkboxes)
            checkbox_option_text_list = []
            geo_list_smstest = []
            failed_call_data = {}
            driver.implicitly_wait(1)
            for test in tests:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        checkbox_option_text_list = values
                        break
                    else:
                        checkbox_option_text_list = []
                try:
                    operator_comparsion_with_by_reading_pdf_export_file(driver,checkbox_option_text_list,enabled_checkboxes,disabled_checkboxes,pdf_view.parent_checkbox_pdf,Title,test,geo_list_smstest,failed_call_data,result_status,data_difference,data_same,excelpath,excelpath_for_storedata,remote_test_campaign,campaign,device)
                except Exception as e:
                    continue
            try:
                if len(failed_call_data) !=0:
                    try:
                        # Initialize a dictionary to store the total values for 'Total Geo samples'
                        total_geo_samples = 0
                        # Iterate through the dictionary to calculate the total values
                        for key, value_list in failed_call_data.items():
                            for i in range(len(value_list)):
                                if len(value_list[i]) > 1 and str(value_list[i][0]).replace(" ","").lower() == str('Total Geo samples').replace(" ","").lower():
                                    total_geo_samples += int(value_list[i][1])

                        for key, value_list in failed_call_data.items():
                            for i in range(len(value_list)):
                                if len(value_list[i]) > 1 and str(value_list[i][0]).replace(" ","").lower() == str('Total Geo samples').replace(" ","").lower():
                                    value_list[i] = [str('Total Geo samples').replace(" ","").lower(), str(total_geo_samples)]
                    except:
                        pass
                    for checkbox_option_text_list3 ,Export_pdf_table_data in failed_call_data.items():
                        try:
                            checkbox_option_text_list = [checkbox_option_text_list3]
                            data_comparison_in_pdf_export(driver,Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest,result_status,data_difference,data_same,excelpath,excelpath_for_storedata,remote_test_campaign,campaign,device)
                        except:
                            continue
            except:
               pass
            try:
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            except:
                pass
            time.sleep(5)
            if "Default Settings" == settingvalue:
                g_flag = 0
                try:
                    time.sleep(2)
                    clickec(driver,pdf_view.save_pdf_export)
                    with allure.step("Screenshot to verify clicked on generate pdf"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"generate_pdf_screenshot",attachment_type=allure.attachment_type.PNG)
                    g_flag = 1
                except Exception as e:
                    raise e
                if g_flag == 1:
                    try:
                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                        # Set the start time of the loop
                        try:
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 60
                            if len(Generating_report) == 0:
                                with allure.step("Waiting for generate pdf to load"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for generate pdf to load",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                        # Check if the condition is met
                                        if len(Generating_report) != 0:
                                           break
                        except:
                            pass
                        try:
                            pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                        except:
                            pass
                        try:
                            Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 1800
                            if len(Generating_report) != 0:
                                with allure.step("Waiting for generate pdf for downloading"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for generate pdf for downloading",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                        # Check if the condition is met
                                        if len(Generating_report) == 0:
                                           break
                        except:
                            pass
                        try:
                            pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                            start_time = time.time()
                            # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                            max_run_time = 60
                            if len(pdf_files) == 0:
                                with allure.step("Waiting for complete pdf download"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for complete pdf download",attachment_type=allure.attachment_type.PNG)
                                    while time.time() - start_time < max_run_time:
                                        pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                                        # Check if the condition is met
                                        if len(pdf_files) != 0:
                                           break
                        except:
                            pass
                        time.sleep(2)
                        pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                        if len(pdf_files) == 0:
                            statment ="failed step:- pdf is not downloaded"
                            with allure.step("failed step:- PDF is not downloaded"):
                                allure.attach(driver.get_screenshot_as_png(), name=f"failed step:- PDF is not downloaded",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus2 = status(Title,"pdf download", "FAILED", statment)
                                result_status.put(updatecomponentstatus2)
                                e = Exception
                                raise e
                        elif len(pdf_files) != 0:
                            statment = "pdf is downloaded successfully"
                            with allure.step("PDF is downloaded successfully"):
                                allure.attach(driver.get_screenshot_as_png(),name=f"PDF is downloaded successfully",attachment_type=allure.attachment_type.PNG)
                                updatecomponentstatus2 = status(Title, "pdf download", "PASSED",statment)
                                result_status.put(updatecomponentstatus2)
                    except Exception as e:
                        pass
                elif g_flag == 0:
                    statment = f"failed step:- failed to click on save as pdf btn"
                    with allure.step(f"failed step:- failed to click on save as pdf btn"):
                        updatecomponentstatus2 = status(Title, "pdf download", "FAILED", statment)
                        result_status.put(updatecomponentstatus2)
                        e = Exception
                        raise e
    except Exception as e:
        pass
    finally:
        try:
            driver.close()
        except Exception as e:
            pass
        try:
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass
def update_result_of_pdf(result_status,data_difference,data_same,excel_file_path,data_match_sheet,data_not_match_sheet):
    try:
        status = []
        df_data_difference = []
        df_data_same = []
        combined_status_df = None
        combined_data_differences = None
        combined_data_same = None
        while not result_status.empty():
            updatecomponentstatus2 = result_status.get()
            df = pd.DataFrame(updatecomponentstatus2)
            status.append(df)
        while not data_same.empty():
            datasame = data_same.get()
            df_same = pd.DataFrame(datasame)
            df_data_same.append(df_same)
        while not data_difference.empty():
            datadiffernce= data_difference.get()
            df_difference = pd.DataFrame(datadiffernce)
            df_data_difference.append(df_difference)
        if len(status) != 0:
            combined_status_df = pd.concat(status, ignore_index=True)
        if len(df_data_difference) !=0:
            combined_data_differences = pd.concat(df_data_difference, ignore_index=True)
        if len(df_data_same) !=0:
            combined_data_same = pd.concat(df_data_same, ignore_index=True)
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet_componentstatus = workbook["COMPONENTSTATUS"]
        worksheet_data_match = workbook[data_match_sheet]
        worksheet_data_not_match = workbook[data_not_match_sheet]
        if len(status) != 0:
            with allure.step("update_result_of_pdf"):
                update_component_status_openpyxl(worksheet=worksheet_componentstatus, dataframe=combined_status_df)
        if len(df_data_difference) !=0:
            with allure.step("combined_data_differences"):
                update_excel_datavalidation_pdf_each_testcase_openpyxl(df=combined_data_differences,worksheet=worksheet_data_not_match)
        if len(df_data_same) !=0:
            with allure.step("combined_data_same"):
                update_excel_datavalidation_pdf_each_testcase_openpyxl(df=combined_data_same,worksheet=worksheet_data_match)
        workbook.save(excel_file_path)
        workbook.close()
    except Exception as e:
        pass
def process_enabled_checkbox(Title,result_status,enabled_checkboxes,enabled_checkboxs,disabled_checkboxes):
    if any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "PASSED","In the 'TC' sheet of testdata.xlsx, a particular component is marked as 'yes' and the checkbox associated with it is enabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx, for a particular component, the checkbox is marked as 'yes,' but it is disabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(enabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(enabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, enabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx, the 'yes' value for a particular component does not correspond to the enabled/disabled checkbox options")
        result_status.put(updatecomponentstatus2)
def process_disabled_checkbox(Title,result_status,disabled_checkboxs,enabled_checkboxes,disabled_checkboxes):
    if not any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "PASSED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and checkbox for that component is disabled.,")
        result_status.put(updatecomponentstatus2)
    elif any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and checkbox for that component is enabled")
        result_status.put(updatecomponentstatus2)
    elif not any(re.fullmatch(disabled_checkboxs[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(disabled_checkboxs[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
        updatecomponentstatus2 = status(Title, disabled_checkboxs[0].strip(), "FAILED","In the 'TC' sheet of testdata.xlsx,particular component's cell is empty and does not correspond to the enabled/disabled checkbox options")
        result_status.put(updatecomponentstatus2)

def operator_comparsion_with_by_reading_pdf_export_file(driver,checkbox_option_text_list,enabled_checkboxes,disabled_checkboxes,parent_pdf_export_checkbox,Title, test,geo_list_smstest,failed_call_data,result_status,data_difference,data_same,excelpath,excelpath_for_storedata,remote_test_campaign,campaign,device):
    Export_pdf_table_data  = "None"
    locator  = "None"
    Export_pdf_table1 = "None"
    Export_pdf_table_data = None
    Export_pdf_table_data1 = []
    try:
        if checkbox_option_text_list.__len__() == 0:
            with allure.step(f"In input data from 'pdf_export'sheet {str(config.pdf_export_excel_path)} for 'PDF EXPORT VIEW for '{test}' in header of PDF EXPORT VIEW Components 1st column value against the 2nd row of headers of PDF EXPORT VIEW in {str(config.test_data_path)} is mismatch/empty"):
                allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                e = Exception
                raise e
        elif checkbox_option_text_list.__len__() != 0:
            try:
                if any(re.fullmatch(checkbox_option_text_list[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and not any(re.fullmatch(checkbox_option_text_list[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
                    with allure.step(f"{Title} of {checkbox_option_text_list[0]} and it is enabled"):
                        try:
                            checkbox_option_text_locator = []
                            try:
                                pattern_mapping_locator_df = pd.read_excel(config.pdf_export_excel_path, sheet_name="pdf_locators_for_table")
                            except Exception as e:
                                with allure.step(f"Check {config.pdf_export_excel_path}"):
                                    print(f"Check {config.pdf_export_excel_path}")
                                    assert False
                            pattern_mapping_locator = pattern_mapping_locator_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(), axis=1).to_dict()
                            for pattern, values in pattern_mapping_locator.items():
                                if pattern.lower() == checkbox_option_text_list[0].lower():
                                    checkbox_option_text_locator = values
                                    break
                                else:
                                    checkbox_option_text_locator = []

                            if len(checkbox_option_text_locator) !=0:
                                Export_pdf_table = pdf_view.Export_pdf_table
                                for checkbox_option_text_locator1 in checkbox_option_text_locator:
                                    for Export_pdf_tablelocator_dict in Export_pdf_table:
                                        Export_pdf_table1 = (Export_pdf_tablelocator_dict['locator by'], Export_pdf_tablelocator_dict['locator'].format(checkbox_option_text_locator1))
                                        try:
                                            Export_pdf_table1_element = driver.find_element(*Export_pdf_table1)
                                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});",Export_pdf_table1_element)
                                        except Exception as e:
                                            pass
                                        Export_pdf_table_data = extract_table_datas_span(driver, Export_pdf_table1,f"{Title} table of {checkbox_option_text_list[-1]}",checkbox_option_text_list[-1], Title,excelpath)
                                        Export_pdf_table_data1.append(Export_pdf_table_data)
                            elif len(checkbox_option_text_locator) ==0 or checkbox_option_text_locator == None:
                                with allure.step(f"In input data from 'pdf_locators_for_table'sheet {str(config.pdf_export_excel_path)} for 'PDF EXPORT VIEW for '{test}' in header of PDF EXPORT VIEW Components 1st column value against the 2nd row of headers of PDF EXPORT VIEW in {str(config.test_data_path)} is mismatch/empty"):
                                    allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                                    e = Exception
                                    raise e
                        except:
                            pass
                        try:
                            pattern_mapping_df = pd.read_excel(config.map_view_components_excelpath,sheet_name="MAPVIEW_PDFVIEW")
                        except Exception as e:
                            with allure.step(f"Check {config.map_view_components_excelpath}"):
                                print(f"Check {config.map_view_components_excelpath}")
                                assert False
                        pattern_mapping = pattern_mapping_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(), axis=1).to_dict()
                        print(pattern_mapping)
                        test = checkbox_option_text_list[-1]
                        option_text_listpdf = []
                        # Remove leading and trailing spaces from test
                        for pattern, values in pattern_mapping.items():
                            if pattern.lower() == test.lower():
                                option_text_listpdf = values
                                break
                            else:
                                option_text_listpdf = []
                        for option_text_listpdf1,Export_pdf_table_data in zip(option_text_listpdf,Export_pdf_table_data1):
                            try:
                                d_flag = 0
                                a_flag = 0
                                try:
                                    if len(Export_pdf_table_data) != 0:
                                        Export_pdf_table_data1 = []
                                        for data in Export_pdf_table_data:
                                            Export_pdf_table_data1.append(data)
                                        Export_pdf_table_data1.insert(0, [checkbox_option_text_list[-1]])
                                        Export_pdf_table_data1.append(["ENDHERE"])
                                        Export_pdf_table_data.insert(0, [option_text_listpdf1])
                                        Export_pdf_table_data.append(["ENDHERE"])
                                        try:
                                            export_pdf_update_to_excel(Export_pdf_table_data1, "PDF_EXPORT",checkbox_option_text_list[-1], excelpath_for_storedata)
                                            a_flag = 1
                                        except Exception as e:
                                            if len(Export_pdf_table_data) != 0 and d_flag == 0 and a_flag == 0:
                                                with allure.step(f"Failed step :- In {Title} table for {checkbox_option_text_list[-1]}/{option_text_listpdf1} error in insert/appending data to Excel report"):
                                                    updatecomponentstatus2 = status(Title,f"{checkbox_option_text_list[-1]}/{option_text_listpdf1}","FAILED",f"Failed step :- In {Title} table for {checkbox_option_text_list[-1]}/{option_text_listpdf1} error in insert/appending data to Excel report")
                                                    result_status.put(updatecomponentstatus2)
                                        try:
                                            if not any(checkbox_option_text_list[-1] == checkboxvalue for checkboxvalue in ["CallDrop", "CallAborted","CallSetupFailure","CallNoNetwork"]):
                                                data_comparison_in_pdf_export(driver, Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest, result_status,data_difference, data_same, excelpath,excelpath_for_storedata,remote_test_campaign,campaign,device)
                                            elif any(checkbox_option_text_list[-1] == checkboxvalue for checkboxvalue in ["CallDrop", "CallAborted","CallSetupFailure","CallNoNetwork"]):
                                                failed_call_data[checkbox_option_text_list[-1]] = Export_pdf_table_data
                                            d_flag = 1
                                        except Exception as e:
                                            raise e
                                    elif len(Export_pdf_table_data) == 0 or Export_pdf_table_data == None:
                                        e = Exception
                                        raise e
                                except Exception:
                                    if len(Export_pdf_table_data) == 0 or Export_pdf_table_data == None:
                                        with allure.step(f"Failed step :- The {Title} for {checkbox_option_text_list[-1]}/{option_text_listpdf1} does not contain any data table."):
                                            updatecomponentstatus2 =status(Title,checkbox_option_text_list[-1],"FAILED",f"Failed step :- The {Title} for {checkbox_option_text_list[-1]} does not contain any data table")
                                            result_status.put(updatecomponentstatus2)
                                            raise Exception
                            except Exception as e:
                                continue
                elif not any(re.fullmatch(checkbox_option_text_list[0].strip(), enabled_checkbox.strip(), re.IGNORECASE) for enabled_checkbox in enabled_checkboxes) and any(re.fullmatch(checkbox_option_text_list[0].strip(), disabled_checkbox.strip(), re.IGNORECASE) for disabled_checkbox in disabled_checkboxes):
                    with allure.step(f"Failed step :- {checkbox_option_text_list[-1]} checkbox is disabled"):
                        raise Exception
            except Exception as e:
                raise e
    except Exception as e:
        raise e
def data_comparison_in_pdf_export(driver,Export_pdf_table_data,checkbox_option_text_list,geo_list_smstest,result_status,data_difference,data_same,excelpath,excelpath_for_storedata,remote_test_campaign,campaign,device):
    c_flag= None
    s_flag= None
    Title = None
    try:
        Title = "Map view(NQC-operator comparison V/S PDF Export)"
        operator_comparisonsheet = pd.read_excel(excelpath_for_storedata, sheet_name="OPERATOR_COMPARISON")
        start_row_index = None
        # Remove spaces from strings in operator_comparisonsheet
        for i, row in operator_comparisonsheet.iterrows():
            for col in row.index:
                cell_value = str(row[col]).replace(' ', '')
                if any(re.fullmatch(cell_value.strip(),Export_pdf_enabled_checkbox_name.replace(' ', '').lower().strip(),re.IGNORECASE) for Export_pdf_enabled_checkbox_name in Export_pdf_table_data[0]):
                    start_row_index = i
                    break
            if start_row_index != None:
                break
        print(start_row_index)
        end_row_index = 0
        s_flag = 0
        if start_row_index != None:
            s_flag = 1
            for i, row in operator_comparisonsheet.iterrows():
                if i >= start_row_index:
                    for col in row.index:
                        cell_value = str(row[col]).replace(' ', '')
                        if re.fullmatch(cell_value.strip(), "ENDHERE", re.IGNORECASE):
                            end_row_index = i
                            break
                    if end_row_index != 0:
                        break
            print(end_row_index)
            # Select the desired rows based on the indices
            selected_rows = operator_comparisonsheet.iloc[start_row_index:end_row_index + 1]
            # Optional: Reset the index of the selected rows
            selected_rows = selected_rows.reset_index(drop=True)
            selected_rows_list = selected_rows.values.tolist()
            # Display the selected rows
            if len(selected_rows_list)!=0:
                # Clean Export_pdf_table_data
                Export_pdf_table_data_cleaned = [[str(item).replace(' ', '').lower() for item in sublist if item is not None and item != '' and not pd.isna(item)] for sublist in Export_pdf_table_data]
                # Clean selected_rows_list
                table_data = [[str(item).replace(' ', '').lower() for item in sublistdata if item is not None and not pd.isna(item)] for sublistdata in selected_rows_list]
                try:
                    if checkbox_option_text_list[-1] == "SmsSent" or checkbox_option_text_list[-1] == "SmsRecieve":
                        # Use list comprehension to find sublists containing the word 'Geo'
                        geo_list = [sublist for sublist in table_data if any('Geo'.lower() in str(item).strip().lower() for item in sublist)]
                        for geo in geo_list:
                            geo_list_smstest.append(geo)
                except:
                    pass
                if checkbox_option_text_list[-1].lower() == "SmsTest".lower():
                    # Create a dictionary to store the sum of values based on the "Total Geo samples" key
                    geo_sum_dict = {}
                    for item in geo_list_smstest:
                        key, value = item[0], item[1]
                        if key in geo_sum_dict:
                            geo_sum_dict[key] += int(value)
                        else:
                            geo_sum_dict[key] = int(value)
                    # Convert the dictionary back to a list of sublists
                    combined_geo_list = [[key.replace(" ", ""), str(value)] for key, value in geo_sum_dict.items()]
                    # Find the index of the sublist containing the word "Geo"
                    geo_index = next((i for i, sublist in enumerate(table_data) if any('geo' in item.lower() for item in sublist)),None)
                    if geo_index is not None:
                        # Replace the sublist with the combined_geo_list
                        table_data[geo_index] = combined_geo_list[0]
                if len(table_data) != 0:
                    # Compare the lists element-wise and find the differences
                    differences = []
                    similar = []
                    datas = []
                    c_flag = 0
                    if len(table_data) == len(Export_pdf_table_data_cleaned):
                        for i, (row_data, row_export) in enumerate(zip(table_data, Export_pdf_table_data_cleaned)):
                            for j, (data_item, export_item) in enumerate(zip(row_data, row_export)):
                                if not compare_values(data_item.lower(),export_item.lower()):
                                    differences.append(f"Difference at position ({i}, {j}): {data_item} vs {export_item}")
                                elif compare_values(data_item.lower(),export_item.lower()):
                                    similar.append(f"Same at position ({i}, {j}): {data_item} vs {export_item}")
                    elif len(table_data) != len(Export_pdf_table_data_cleaned):
                        # Create an empty dictionary
                        operator_table_data_dict1 = {}
                        operator_table_data_dict = []
                        # Loop through the data list
                        for sublist in table_data:
                            # Check if the sublist has at least two items
                            if len(sublist) >= 2:
                                key = sublist[0]
                                value = sublist[1]
                                operator_table_data_dict1[key] = value
                        operator_table_data_dict.append(operator_table_data_dict1)
                        Export_pdf_table_data_dict1 = {}
                        Export_pdf_table_data_dict = []
                        # Loop through the data list
                        for sublist in Export_pdf_table_data_cleaned:
                            # Check if the sublist has at least two items
                            if len(sublist) == 2:
                                key = sublist[0]
                                value = sublist[1]
                                Export_pdf_table_data_dict1[key] = value
                        Export_pdf_table_data_dict.append(Export_pdf_table_data_dict1)
                        for Export_pdf_table_data_item, operator_table_data_item in zip(Export_pdf_table_data_dict, operator_table_data_dict):
                            for key in Export_pdf_table_data_item:
                                try:
                                    if not compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        differences.append(f"Difference in key value '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                                    elif compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        similar.append(f"Same in key value '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                                except Exception as e:
                                    if not compare_values( Export_pdf_table_data_item[key],operator_table_data_item.get(key, 'Key_name_cant_find')):
                                        differences.append(f"Key name can't find in operator comparsion '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, 'Key_name_cant_find')}")
                                    elif compare_values( Export_pdf_table_data_item[key],operator_table_data_item[key.strip()]):
                                        similar.append(f"Same in key '{key}': {Export_pdf_table_data_item[key]} vs {operator_table_data_item.get(key, '')}")
                    c_flag = 1
                    df_differences = None
                    if differences:
                        if len(similar) != 0:
                            df_differences = {
                                'Usercampaignname,Classifier,Device': [f"{remote_test_campaign},{campaign},{device}"]* (len(similar) + len(differences)+2),
                                'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(similar) + len(differences)+2),
                                'Data validation': ["STARTHERE"] + similar + differences + ["ENDHERE"]
                            }
                        elif len(similar) == 0:
                            df_differences = {
                                'Usercampaignname,Classifier,Device': [f"{remote_test_campaign},{campaign},{device}"] * (len(similar) + len(differences) + 2),
                                'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(differences)+2),
                                'Data validation': ["STARTHERE"] + differences + ["ENDHERE"]
                            }
                        if df_differences != None:
                            data_difference.put(df_differences)
                        statement = f"Failed step:- There is difference in data when comparing map_view against pdf_view ."
                        with allure.step(statement+f" for {checkbox_option_text_list[-1]}"):
                            html_for_csv(datas,checkbox_option_text_list[-1])
                            updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED", statement,)
                            result_status.put(updatecomponentstatus2)
                            raise Exception
                    elif len(similar) != 0 and len(differences) == 0:
                        df_similar = {
                            'Usercampaignname,Classifier,Device': [f"{remote_test_campaign},{campaign},{device}"] * (len(similar)+ 2),
                            'Component Type': [f'Component {checkbox_option_text_list[-1]}'] * (len(similar)+2),
                            'Data validation': ["STARTHERE"] + similar + ["ENDHERE"]
                        }
                        data_same.put(df_similar)
                        statement = f"There is same data when comparing map_view against pdf_view"
                        with allure.step(statement + f"for {checkbox_option_text_list[-1]}"):
                            updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "PASSED", statement)
                            result_status.put(updatecomponentstatus2)
                            html_for_csv(datas,checkbox_option_text_list[-1])
        elif start_row_index == None:
            statement = f"Failed step :- There is no data found in the Operator comparison sheet for this component as a reference."
            with allure.step(statement + f":- {checkbox_option_text_list[-1]}"):
                updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED",statement)
                result_status.put(updatecomponentstatus2)
                raise Exception
    except Exception as e:
        if c_flag == 0 and s_flag == 1:
            statement = f"Failed step :- error in comparing data, Please check the testcase excel report of 'PDF_EXPORT' sheet and locators path in 'pdf_locators_for_table' sheet for"
            with allure.step(statement + f":- {checkbox_option_text_list[-1]}"):
                updatecomponentstatus2 = status(Title, checkbox_option_text_list[-1], "FAILED", statement)
                result_status.put(updatecomponentstatus2)
                raise Exception
        # Handle exceptions
        print(f"Error: {e}")
        raise e

def update_excel_datavalidation_pdf_each_testcase_openpyxl(df,worksheet):
    """
        Update the high-level Excel report for data validation of PDF.
        Args:
            df (DataFrame): The DataFrame containing data to be added to the Excel sheet.
            sheet (Excel Sheet): The Excel sheet to update with the data and formatting.
        Returns:
            None
        """
    try:
        color_mapping = {
            'STARTHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            'ENDHERE': PatternFill(start_color="D2B48C", end_color="D2B48C", fill_type="solid"),  # Light Brown
            "Same": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),  # Green
            "Difference": PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid"),  # Red
            "Key name can't find in operator comparsion": PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid") ,  # Yellow
            "Key not present in combine_binary_export": PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type="solid")  # Light Yellow
        }
        # Find the last used row in the sheet
        last_row = worksheet.max_row
        start_row = last_row + 1
        # Insert the DataFrame into the worksheet
        for index, row in df.iterrows():
            worksheet.append(row.tolist())
        # Apply color formatting to the entire range
        for i, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=start_row + len(df) - 1, min_col=2, max_col=2),start=start_row):
            validation_cell = row[0]
            data_validation = df.iloc[i - start_row]["Data validation"]
            for keyword, fill in color_mapping.items():
                if keyword in data_validation:
                    validation_cell.fill = fill
        # Set colors for File and ParameterType columns
        for i in range(start_row, start_row + len(df)):
            worksheet.cell(row=i, column=1).fill = PatternFill(start_color='FFC864', end_color='FFC864',fill_type="solid")  # Light Orange for ParameterType column
    except Exception as e:
        with allure.step(f"{str(e)}"):
            pass

##########################################################################################
def floor_plan_for_individual_campaign(driver, userid, password, excelpath, test_case_downloading_files_path):
    status_floor_plan_map_flag = True
    floor_plan_map_title = "Floor Plan"
    Floorplan_runvalue = floorplan_module_controllers()
    if "Yes".lower() == Floorplan_runvalue[-1].strip().lower():
        campaigns_datas = fetch_camapaigns(sheet_to_run=["Floor Plan Data"])
        driver.implicitly_wait(3)
        for campaigns_data in campaigns_datas:
            device, campaign, usercampaignsname, testgroup = campaigns_data

            with allure.step(f"Usercampaignname:{usercampaignsname},Classifier:{campaign},Device:{device} --> Floor Plan"):

                downloadfilespath = specifying_download_path(driver, test_case_downloading_files_path + "\\",campaign + "_" + usercampaignsname)

                protestdata_runvalue = protestdata_module_controllers()
                litetestdata_runvalue = litetestdata_module_controllers()
                typeoftest = None

                if "Yes".lower() == protestdata_runvalue[-1].strip().lower():
                    typeoftest = "ProTest data"
                elif "Yes".lower() == litetestdata_runvalue[-1].strip().lower():
                    typeoftest = "LiteTest data"

                with allure.step(f"Navigating to [Android TestData  >>> {typeoftest}  >>>  Device  >>>  {str(campaign)}]"):
                    side_menu_Components_(driver, device, usercampaignsname, userid, password, excelpath)

                with allure.step("Floor Plan"):
                    status_floor_plan_map,floor_plan_map_title = Floor_plan(driver,campaign,downloadfilespath,excelpath)
                    if status_floor_plan_map == "FAILED":
                        status_floor_plan_map_flag = False
        driver.implicitly_wait(30)
        update_high_level_status_of_floor_plan_map(status_floor_plan_map_flag, Title=floor_plan_map_title, excelpath=excelpath)
        # update_module_status_based_on_reading_component_status(modules=[f"Floor Plan"], excelpath=excelpath,condition="contains")
    elif "Yes".lower() != Floorplan_runvalue[-1].strip().lower():
        updatehighmodulestatus(Title=floor_plan_map_title, status="SKIPPED", comments=f"You have selected No for execute", path=excelpath)
def Floor_plan(driver,campaign,downloadfilespath,excelpath):
    Title = "Floor Plan"
    floor_plan_map_title = f"{Title}-->MAP VIEW"
    floor_plan_pdf_title = f"{Title}-->PDF VIEW"
    try:
        result_status = queue.Queue()
        remote_test_point, map_start_point, graph_start_point, export_start_point, load_start_point, PDF_Export_index_start_point, END_index = fetch_input_points()
        tests = fetch_components(campaign, map_start_point, graph_start_point)
        tests_pdf = fetch_components(campaign, PDF_Export_index_start_point, END_index)
        tests_no_pdf = fetch_components_for_no_yes(campaign, PDF_Export_index_start_point, END_index)
        Map_view(driver, tests, excelpath, "Floor Plan",run_sub_modules=["Floor Plan"])
        # pdf_comparsion_for_floor_plan(driver, tests_pdf, floor_plan_pdf_title, result_status, tests_no_pdf, downloadfilespath,floorplan.floormap,Title, excelpath)
    except Exception as e:
        print(e)
    finally:
        floor_map_and_pdf_title = floor_plan_map_title +"|"+floor_plan_pdf_title
        status = readcomponentstatus_(status_word="WARNING|FAILED", path=excelpath, Titlename=floor_map_and_pdf_title,condition="contains")
        if status == "FAILED":
            return "FAILED" , floor_plan_map_title
        elif status == "PASSED":
            return "PASSED" , floor_plan_map_title
def update_high_level_status_of_floor_plan_map(status_floor_plan_map_flag,Title,excelpath):
    if status_floor_plan_map_flag == True:
        updatehighmodulestatus(Title, status="PASSED", comments=f"{Title} is Passed", path=excelpath)
    elif status_floor_plan_map_flag == False:
        updatehighmodulestatus(Title, status="FAILED", comments=f"{Title} is failed", path=excelpath)


def pdf_comparsion_for_floor_plan(driver,tests,Title,result_status,tests_no,floormap,elementname,downloadpdfpath,excelpath):
    enabled_checkboxes = None
    disabled_checkboxes = None
    checkbox_option_text_list = None
    pdf_files  = None
    try:
        List_of_options_txt = ["Export As PDF"]
        # Load pattern mapping from Excel file
        try:
            pattern_mapping_df = pd.read_excel(config.pdf_export_excel_path,sheet_name="pdf_components")
        except Exception as e:
            with allure.step(f"Check {config.map_view_components_excelpath}"):
                print(f"Check {config.map_view_components_excelpath}")
                assert False
        flag, alert_text = select_from_listbox_ECs(driver, List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown,List_Of_Campaigns_Export_Dashboard.List_Of_Campaigns_Export_Dropdown_Options,List_of_options_txt, Title, excelpath)
        # Convert pattern mapping to dictionary
        pattern_mapping = pattern_mapping_df.set_index('pdf_export').apply(lambda x: x.dropna().tolist(),axis=1).to_dict()
        # Match patterns with tests
        txt = []
        if tests.__len__() == 0:
            statement = f"{Title}  --  Nothing is marked 'Yes' in {str(config.test_data_path)}"
            with allure.step(f"Nothing is marked 'Yes' in {str(config.test_data_path)} for {Title}"):
                updatecomponentstatus2 = status(Title, "", "FAILED",f"Nothing marked in {str(config.test_data_path)}")
                result_status.put(updatecomponentstatus2)
                e = Exception
                raise e
        else:
            enabled_txt=[]
            txts = []
            for test in tests:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        txts = values
                        enabled_txt.append(txts)
                        break
                    else:
                        txts = []
            disabled_txt = []
            disabledtxts = []
            for test in tests_no:
                test = test.strip()  # Remove leading and trailing spaces from test
                for pattern, values in pattern_mapping.items():
                    if pattern.lower() == test.lower():
                        disabledtxts = values
                        disabled_txt.append(disabledtxts)
                        break
                    else:
                        disabledtxts = []
            s_flag = 0
            try:
                driver.switch_to.window(driver.window_handles[1])
                s_flag = 1
            except Exception as e:
                pass
            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
            try:
                try:
                    driver.switch_to.window(driver.window_handles[1])
                except Exception as e:
                    pass
                start_time = time.time()
                # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                max_run_time = 60
                if len(pdf_export_checkbox) == 0:
                    with allure.step("Waiting for pdf page to load"):
                        allure.attach(driver.get_screenshot_as_png(), name=f"Waiting for pdf page to load",attachment_type=allure.attachment_type.PNG)
                        while time.time() - start_time < max_run_time:
                            try:
                                driver.switch_to.window(driver.window_handles[1])
                            except Exception as e:
                                pass
                            pdf_export_checkbox = driver.find_elements(*pdf_view.parent_checkbox_pdf)
                            # Check if the condition is met
                            if len(pdf_export_checkbox) != 0:
                                break
            except:
                pass
            time.sleep(5)
            enabled_checkboxes, disabled_checkboxes = check_selected_and_finding_enable_and_disabled_checkboxes_(driver,pdf_view.parent_checkbox_pdf)
            with concurrent.futures.ThreadPoolExecutor() as executor:
                for enabled_checkboxs in enabled_txt:
                    executor.submit(process_enabled_checkbox, Title, result_status, enabled_checkboxes,enabled_checkboxs, disabled_checkboxes)
            with concurrent.futures.ThreadPoolExecutor() as executor:
                for disabled_checkboxs in disabled_txt:
                    executor.submit(process_disabled_checkbox, Title, result_status, disabled_checkboxs,enabled_checkboxes, disabled_checkboxes)

            floorplan_btn_element = driver.find_element(floorplan.show_all_floor_plan_toggle_brn_in_pdf_window[0], floorplan.show_all_floor_plan_toggle_brn_in_pdf_window[1])
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", floorplan_btn_element)
            clickec(driver, floorplan.show_all_floor_plan_toggle_brn_in_pdf_window)

            try:
                for test in tests:
                    test = test.strip()  # Remove leading and trailing spaces from test
                    for pattern, values in pattern_mapping.items():
                        if pattern.lower() == test.lower():
                            checkbox_option_text_list = values
                            break
                        else:
                            checkbox_option_text_list = []
                    if checkbox_option_text_list != []:
                        # Read the Excel file
                        id_testtypes = pd.read_excel(config.pdf_export_excel_path, sheet_name="floorplan_pdf")

                        # Create a dictionary from the Excel columns
                        id_testtypes_dict = id_testtypes.set_index('pdf_export')[['Locators1', 'Locators2']].apply(lambda x: [i for i in x if pd.notna(i)], axis=1).to_dict()

                        try:
                            # Get the test type locators from the dictionary
                            id_testtype = id_testtypes_dict.get(checkbox_option_text_list[0], [])
                            for id in id_testtype:
                                try:
                                    floorplan_map_for_pdf = driver.find_elements(By.XPATH,f"//div[@id='{id}']//img[@src = 'https://maps.gstatic.com/mapfiles/transparent.png']/parent::div[contains(@style,'width: 64px; height: 64px;')]")
                                    try:
                                        for floorplan_maps in floorplan_map_for_pdf:
                                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});",floorplan_maps)
                                    except:
                                        pass

                                    if any(floorplan_maps.is_displayed() for floorplan_maps in floorplan_map_for_pdf):
                                        updatecomponentstatus(Title, str(test), "PASSED", "Floor Plan found", excelpath)
                                        allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                                    elif not any(floorplan_maps.is_displayed() for floorplan_maps in floorplan_map_for_pdf):
                                        updatecomponentstatus(Title, str(test), "FAILED", "Floor Plan not found",excelpath)
                                        allure.attach(driver.get_screenshot_as_png(), name=f"{test}_screenshot",attachment_type=allure.attachment_type.PNG)
                                except Exception as e:
                                    updatecomponentstatus(Title, str(test), "WARNING", f"Should be handled {e}",excelpath)
                                    pass
                        except Exception as e:
                            pass
                    elif checkbox_option_text_list == []:
                        updatecomponentstatus(Title, str(test), "WARNING", f"Should be handled with name", excelpath)
                        pass
            except Exception as e:
                pass  # Continue processing even
            try:
                time.sleep(2)
                clickec(driver, pdf_view.save_pdf_export)
                with allure.step("Screenshot to verify clicked on generate pdf"):
                    allure.attach(driver.get_screenshot_as_png(), name=f"generate_pdf_screenshot",attachment_type=allure.attachment_type.PNG)

            except Exception as e:
                raise e

            try:
                Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                # Set the start time of the loop
                try:
                    start_time = time.time()
                    # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                    max_run_time = 60
                    if len(Generating_report) == 0:
                        with allure.step("Waiting for generate pdf to load"):
                            allure.attach(driver.get_screenshot_as_png(),name=f"Waiting for generate pdf to load",attachment_type=allure.attachment_type.PNG)
                            while time.time() - start_time < max_run_time:
                                Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                # Check if the condition is met
                                if len(Generating_report) != 0:
                                    break
                except:
                    pass
                try:
                    pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                except:
                    pass
                try:
                    Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                    start_time = time.time()
                    # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                    max_run_time = 1800
                    if len(Generating_report) != 0:
                        with allure.step("Waiting for generate pdf for downloading"):
                            allure.attach(driver.get_screenshot_as_png(),name=f"Waiting for generate pdf for downloading",attachment_type=allure.attachment_type.PNG)
                            while time.time() - start_time < max_run_time:
                                Generating_report = driver.find_elements(*pdf_view.generate_report_pdf)
                                # Check if the condition is met
                                if len(Generating_report) == 0:
                                    break
                except:
                    pass
                try:
                    pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                    start_time = time.time()
                    # Maximum time in seconds the loop should run (1 minute = 60 seconds)
                    max_run_time = 60
                    if len(pdf_files) == 0:
                        with allure.step("Waiting for complete pdf download"):
                            allure.attach(driver.get_screenshot_as_png(),
                                          name=f"Waiting for complete pdf download",
                                          attachment_type=allure.attachment_type.PNG)
                            while time.time() - start_time < max_run_time:
                                pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                                # Check if the condition is met
                                if len(pdf_files) != 0:
                                    break
                except:
                    pass
                time.sleep(2)
                pdf_files = glob.glob(f'{downloadpdfpath}/*.pdf')
                if len(pdf_files) == 0:
                    statment = "failed step:- pdf is not downloaded"
                    with allure.step("failed step:- PDF is not downloaded"):
                        allure.attach(driver.get_screenshot_as_png(),
                                      name=f"failed step:- PDF is not downloaded",
                                      attachment_type=allure.attachment_type.PNG)
                        updatecomponentstatus2 = status(Title, "pdf download", "FAILED", statment)
                        result_status.put(updatecomponentstatus2)
                        e = Exception
                        raise e
                elif len(pdf_files) != 0:
                    statment = "pdf is downloaded successfully"
                    with allure.step("PDF is downloaded successfully"):
                        allure.attach(driver.get_screenshot_as_png(),
                                      name=f"PDF is downloaded successfully",
                                      attachment_type=allure.attachment_type.PNG)
                        updatecomponentstatus2 = status(Title, "pdf download", "PASSED", statment)
                        result_status.put(updatecomponentstatus2)
            except Exception as e:
                pass
                statment = f"failed step:- failed to click on save as pdf btn"
                with allure.step(f"failed step:- failed to click on save as pdf btn"):
                    updatecomponentstatus2 = status(Title, "pdf download", "FAILED", statment)
                    result_status.put(updatecomponentstatus2)
                    e = Exception
                    raise e
                driver.switch_to.window(driver.window_handles[0])

        try:
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass
        pass
    finally:
        try:
            driver.switch_to.window(driver.window_handles[0])
        except:
            pass

############################################################## account icon ################################################################################################
def click_on_account_icon_dropdown_btn(driver):
    clickec(driver,Login_Logout.dropdown_dropdown_toggle)

############################################################## account setting btn in account icon ####################################################################################

def click_on_account_setting_btn(driver):
    clickec(driver,account_icon.account_settings_option)

##################################################################################################################################################

################################################# side bar menu for clicking on list of campaigns used for "Chart" module ###############################################################################################################################################################################################################
def side_bar_menu_for_work_list_campaigns(driver, userid, password, campaigns_datas, excelpath,campaigns_created,device):
    global campaign1, campaign2
    for i, j in zip_longest(range(len(campaigns_created)), range(len(campaigns_datas))):
        device_testdata, campaign, usercampaignsname_testdata, testgroup = campaigns_datas[j]
        remote_test_campaign = campaigns_created[i]
        if compare_values(remote_test_campaign, 'None'):
            campaign1 = campaign
            campaign2 = campaign
        elif not compare_values(remote_test_campaign, 'None'):
            campaign1 = remote_test_campaign
            campaign2 = remote_test_campaign + campaign
        # print(i, campaigns_datas[i])
        Variable_MobileDevice_Xpath = (By.XPATH, f"//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']", str(device))
        protestdata_runvalue = protestdata_module_controllers()
        litetestdata_runvalue = litetestdata_module_controllers()
        if protestdata_runvalue[0].lower() == 'Yes'.lower() and litetestdata_runvalue[0].lower() == 'No'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH,f"//span[text()='Pro TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']",str(device))
        elif protestdata_runvalue[0].lower() == 'No'.lower() and litetestdata_runvalue[0].lower() == 'Yes'.lower():
            Variable_MobileDevice_Xpath = (By.XPATH, f"//span[text()='LITE TestData']/parent::a/following-sibling::ul//a[normalize-space()='{str(device)}']//i[@class='fa fa-angle-left pull-right']",str(device))
        classifier = (By.XPATH, "//span[normalize-space()='" + str(campaign1) + "']", str(campaign2))
        active_element = (By.XPATH,f"//ul[@class='treeview-menu style-1 menu-open']//li[@class='treeview ng-scope active']//a[normalize-space()='{str(device)}']")
        if i == 0:
            click_on_side_bar_menu_compnents(driver, device, userid, password, Variable_MobileDevice_Xpath, classifier)
            count_find_the_campaign = searching_visibility_of_campaigns_by_driver_refresh(driver, classifier, device,userid, password,Variable_MobileDevice_Xpath)
        elif i > 0:
            time.sleep(5)
            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0], Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
            except:
                pass
            click_until_visible_or_active(driver, Variable_MobileDevice_Xpath, active_element)
            try:
                wait_for_loading_elements(driver)
            except:
                pass
            search_campaigns(driver, classifier)
        try:
            result = clickec(driver, classifier)
            time.sleep(3)
            updatecomponentstatus("Side bar menu", f"campaign:- {campaign}, usercampaignsname:- {campaign1}, device:- {device}","PASSED", f"Successfully Selected.", excelpath)
            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0],Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].click();", MobileDevice_element)
            except:
                pass

            if result == False:
                e = Exception
                raise e
        except Exception as e:
            updatecomponentstatus("Side bar menu",f"campaign:- {campaign}, usercampaignsname:- {campaign1}, device:- {device}","FAILED", f"Failed to Select and check the campaign is present in the device",excelpath)
            format_workbook(excelpath)

def click_until_visible_or_active(driver, Variable_MobileDevice_Xpath, active_element_xpath):
    while True:
        try:
            time.sleep(1)

            try:
                MobileDevice_element = driver.find_element(Variable_MobileDevice_Xpath[0],Variable_MobileDevice_Xpath[1])
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", MobileDevice_element)
                driver.execute_script("arguments[0].click();", MobileDevice_element)

            except:pass
            # Wait for the active element to be visible or active
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located(active_element_xpath))
            # Once the active element is visible or active, break out of the loop
            break
        except Exception as e:
            print("Exception occurred:", e)

##########################################################################################################################################################################################################################################################