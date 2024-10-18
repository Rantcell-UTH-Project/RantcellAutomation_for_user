from openpyxl import Workbook
def create_workbook(path):
    """
        Create a new Excel workbook with specific sheet names.
        Args:
            path (str): The path where the new Excel workbook will be saved.
        Notes:
            This function creates a new Excel workbook using openpyxl and adds multiple sheets with predefined names.
            The sheets include 'COMPONENTSTATUS', 'DATAEXTRACT', 'OPERATOR_COMPARISON', 'PDF_EXPORT', 'DATA_MATCH',
            'DATA_NOT_MATCH', 'TABLESUMMARY_DATA_NOT_MATCH', 'TABLESUMMARY_DATA_MATCH', 'CBE_vs_CE_MATCH', and
            'CBE_vs_CE_DONOT_MATCH'.
            After creating the workbook with sheets, it saves the workbook to the specified path.
        """
    workbook = Workbook()
    workbook.create_sheet("HIGH_MODULE_STATUS", 0)
    workbook.create_sheet("COMPONENTSTATUS", 1)
    workbook.create_sheet("RESULTS_DEFAULT_SETTINGS", 2)
    workbook.create_sheet("RESULTS_CHANGE_SETTINGS", 3)
    workbook.create_sheet("Gr_DATA_MATCH", 4)
    workbook.create_sheet("Gr_DATA_NOT_MATCH", 5)
    workbook.create_sheet("OCvspdf_DATAMATCH_defaultsettings", 6)
    workbook.create_sheet("OCvspdf_DATANOTMATCH_defaultsettings", 7)
    workbook.create_sheet("OCvspdf_DATAMATCH_changesettings", 8)
    workbook.create_sheet("OCvspdf_DATANOTMATCH_changesettings", 9)
    workbook.save(path)

def create_workbook_for_data_store(path):
    """
        Create a new Excel workbook with specific sheet names.
        Args:
            path (str): The path where the new Excel workbook will be saved.
        Notes:
            This function creates a new Excel workbook using openpyxl and adds multiple sheets with predefined names.
            The sheets include 'COMPONENTSTATUS', 'DATAEXTRACT', 'OPERATOR_COMPARISON', 'PDF_EXPORT', 'DATA_MATCH',
            'DATA_NOT_MATCH', 'TABLESUMMARY_DATA_NOT_MATCH', 'TABLESUMMARY_DATA_MATCH', 'CBE_vs_CE_MATCH', and
            'CBE_vs_CE_DONOT_MATCH'.
            After creating the workbook with sheets, it saves the workbook to the specified path.
        """
    workbook = Workbook()
    workbook.create_sheet("OPERATOR_COMPARISON", 0)
    workbook.create_sheet("PDF_EXPORT", 1)
    workbook.create_sheet("DATA_EXTRACTION_SETTINGS",2)
    workbook.create_sheet("DATA_EXTRACTION_CHANGE_SETTINGS",3)
    workbook.save(path)

